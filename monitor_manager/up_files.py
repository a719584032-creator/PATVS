from openpyxl import load_workbook
from openpyxl import Workbook
from common.rw_excel import MyExcel
from common.logs import logger
from requests_manager.http_requests_manager import http_manager
import os
import wx
import re


# 检查合并单元格是否符合特定规则
def is_merged_within_range(merged_ranges, row, start_col, end_col):
    for rng in merged_ranges:
        # 检查传入的行号是否在某个合并单元格的范围内，同时列号也要在指定的合并列范围内。
        if row >= rng.min_row and row <= rng.max_row and start_col >= rng.min_col and end_col <= rng.max_col:
            return True
    return False


def get_all_test(file_path, sheet_name):
    """
    获取所有用例，按照指定格式[A-F合并为用例标题，A-D列合并为用例步骤，D-F合并为预期结果]
    """
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    # 获取工作表中所有合并单元格的范围信息，并存入merged_ranges列表。
    merged_ranges = list(sheet.merged_cells.ranges)
    # 获取第19行机型数据
    # row_19_data = [cell.value for cell in sheet[19] if cell.value is not None]
    # model_name = row_19_data[2:]

    # 初始化用例信息存储结构
    case = {
        'title': None,
        'steps': None,
        'expected': None
    }
    all_case = []
    current_title = None
    # 遍历工作表中的所有行
    for row in range(20, sheet.max_row + 1):

        # 检查是否是标题行：A-F列中只有此行数据
        if is_merged_within_range(merged_ranges, row, 1, 6):
            title_cell = sheet.cell(row, 1)
            current_title = title_cell.value
            continue  # 跳过后续步骤，继续下一行

        # 检查是否是步骤：A-D列合并
        if is_merged_within_range(merged_ranges, row, 1, 4):
            steps_cell = sheet.cell(row, 1)
            case['steps'] = steps_cell.value
        # 检查是否是预期结果：E-F列合并
        if is_merged_within_range(merged_ranges, row, 5, 6):
            expected_cell = sheet.cell(row, 5)
            case['expected'] = expected_cell.value
            # 过滤掉都为 None 的数据，添加后重置字典
        if any(case.values()):
            case = {
                'title': current_title if current_title is not None else 'NA',
                'steps': case['steps'] if case['steps'] is not None else 'NA',
                'expected': case['expected'] if case['expected'] is not None else 'NA'
            }
            print('************************************************')
            print(case['expected'])
            print('************************************************')
            all_case.append(case)
            case = {
                'title': None,
                'steps': None,
                'expected': None
            }
    return all_case


# def validate_excel_format(file_name):
#     """
#     校验 TDMS Excel 文件格式
#     """
#     workbook = load_workbook(file_name)
#     required_sheets = ['Plan Information', 'Case List']
#
#     for sheet_name in required_sheets:
#         if sheet_name not in workbook.sheetnames:
#             raise ValueError(f"缺少必要的 sheet 页: {sheet_name}")
#
#     plan_info_sheet = workbook['Plan Information']
#     if plan_info_sheet.cell(1, 1).value != 'Plan name':
#         raise ValueError(
#             f"Plan Information sheet 页中 A1 单元格的值必须为 'Plan name', 实际值为 {plan_info_sheet.cell(1, 1).value}")
#
#     case_list_sheet = workbook['Case List']
#     value = case_list_sheet.cell(1, 2).value
#     value.strip().lower()
#     if value != 'Case name':
#         raise ValueError(
#             f"Case List sheet 页中 B2 单元格的值必须为 'Case name', 实际值为 {value}")
#
#     logger.info("文件格式校验通过")
#     return True

def read_test_cases_from_excel(file_path):
    """
    使用模版上传
    读取Excel文件中的测试用例信息，根据每个Sheet页分别读取用例。
    :param file_path: Excel文件路径
    :return: 字典形式返回，每个键是sheet名称，每个值是该sheet中的测试用例列表
    """
    try:
        workbook = load_workbook(filename=file_path)
        # 初始化一个字典来存储所有 sheet 的用例数据
        all_test_cases = {}
        # 遍历所有 sheet 页
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            # 初始化一个列表来存储当前 sheet 页的所有用例
            test_cases = []
            # 读取当前 sheet 页的内容，从第二行开始读取用例数据
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):  # 跳过空行
                    continue
                test_case = {
                    "model_name": row[0],
                    "title": row[1],
                    "preconditions": row[2],
                    "steps": row[3],
                    "expected": row[4]
                }
                # 将字典加入到当前 sheet 页的用例列表中
                test_cases.append(test_case)
            # 将当前 sheet 页的用例列表加入到总字典中
            all_test_cases[sheet_name] = test_cases
        return all_test_cases
    except FileNotFoundError:
        print(f"文件未找到: {file_path}")
    except Exception as e:
        print(f"读取Excel文件时出错: {e}")


def validate_excel_format(file_path):
    """
    综合校验不同模板的 Excel 文件格式。
    根据文件中的 sheet 页判断模板类型并进行相应的格式校验。

    :param file_path: Excel文件路径
    :return: True 如果校验通过，否则抛出异常
    """
    try:
        # 加载 Excel 文件
        workbook = load_workbook(filename=file_path)

        # 定义两种模板所需的 sheet 页和相应的校验规则
        templates = {
            # TDMS用例模板
            "Template1": {
                "required_sheets": ['Plan Information', 'Case List'],
                "validation_func": validate_template1
            },
            # 自定义模版
            "Template2": {
                "expected_headers": ["测试机型", "用例标题", "前置条件", "用例步骤", "预期结果"],
                "validation_func": validate_template2
            }
        }

        # 判断文件使用的模板
        if set(templates["Template1"]["required_sheets"]).issubset(workbook.sheetnames):
            return templates["Template1"]["validation_func"](workbook)
        else:
            return templates["Template2"]["validation_func"](workbook)

    except FileNotFoundError:
        raise FileNotFoundError(f"文件未找到: {file_path}")
    except Exception as e:
        raise Exception(f"校验Excel文件时出错: {e}")


def validate_template1(workbook):
    """
    校验模板1 TDMS (Plan Information, Case List)
    """
    plan_info_sheet = workbook['Plan Information']
    case_list_sheet = workbook['Case List']

    if plan_info_sheet.cell(1, 1).value != 'Plan name':
        raise ValueError(
            f"Plan Information sheet 页中 A1 单元格的值必须为 'Plan name', 实际值为 {plan_info_sheet.cell(1, 1).value}")

    # 校验 Plan Information 页中单元格是否为空
    if not plan_info_sheet.cell(1, 4).value:
        raise ValueError("Plan Information sheet 页中 project_name 的值不能为空")
    if not plan_info_sheet.cell(4, 4).value:
        raise ValueError("Plan Information sheet 页中 phase 的值不能为空")
    if not plan_info_sheet.cell(8, 2).value:
        raise ValueError("Plan Information sheet 页中 objective 的值不能为空")
    if not extract_values_from_brackets(plan_info_sheet.cell(8, 2).value):
        raise ValueError("Plan Information sheet 页中 objective 的值必须使用[]括起来")
    value = case_list_sheet.cell(1, 2).value
    if value and value.strip().lower() != 'case name':
        raise ValueError(
            f"Case List sheet 页中 B1 单元格的值必须为 'Case name', 实际值为 {value}")

    logger.info("TDMS模板校验通过")
    return 'TDMS'


def validate_template2(workbook):
    """
    校验模板2 power 用例模板 (每个 Sheet 页均符合预期标题格式)
    """
    expected_headers = ["测试机型", "用例标题", "前置条件", "用例步骤", "预期结果"]

    # 遍历每个 Sheet 页
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # 校验标题行
        headers = [cell.value for cell in sheet[1]]
        if headers != expected_headers:
            raise ValueError(f"Sheet '{sheet_name}' 的标题行不符合预期格式: {headers}")

        # 校验用例内容
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # 跳过空行
                continue

            # 校验用例标题、用例步骤和预期结果是否为空
            case_title = row[1]
            case_steps = row[3]
            expected_result = row[4]

            if not case_title or not case_steps or not expected_result:
                raise ValueError(f"Sheet '{sheet_name}' 第 {row_idx} 行: "
                                 f"用例标题、用例步骤或预期结果为空。")

    logger.info("power用例模板校验通过")
    return 'power'


def extract_values_from_brackets(text):
    """
    从给定的字符串中提取所有中括号内的值，返回一个列表。

    :param text: 包含中括号的字符串
    :return: 包含中括号内所有值的列表
    """
    # 使用正则表达式查找所有中括号内的内容
    pattern = r'\[([^\]]+)\]'
    values = re.findall(pattern, text)
    return values


def run_main(file_path, userid, token):
    """
    上传case
    """
    template = validate_excel_format(file_path)
    if template == 'TDMS':
        data = MyExcel(file_path)
        data.active_sheet('Plan Information')
        plan_name = data.get_value_by_rc(1, 2)
        logger.info(f'plan_name is {plan_name}')
        result = http_manager.get_params(f'/get_plan_name_by_planname/{plan_name}/{userid}').get('plan_exists')
        if result:
            raise ValueError(f"当前计划名: {result} 已存在，请勿重复上传")
        project_name = data.get_value_by_rc(1, 4)
        project_phase = data.get_value_by_rc(4, 4)
        # 使用正则表达式查找所有中括号内的内容
        model_names = extract_values_from_brackets(data.get_value_by_rc(8, 2))
        logger.info(f'project_name is {project_name}')
        data.active_sheet('Case List')
        all_sheet = data.getColValues(2)[2:]
        # 实际 sheet_name 需要加上递增的前缀
        all_sheet_with_prefix = [f'{i + 1}-{val}' for i, val in enumerate(all_sheet)]
        all_tester = data.getColValues(14)[2:]
        all_workloading = data.getColValues(15)[2:]
        logger.info(f'all_sheet_with_prefix is {all_sheet_with_prefix}')
        logger.info(f'all_tester is {all_tester}')
        sheet_and_tester_and_workloading = list(zip(all_sheet_with_prefix, all_tester, all_workloading))
        logger.info(f'sheet_and_tester_and_workloading is {sheet_and_tester_and_workloading}')
        for i in sheet_and_tester_and_workloading:
            all_case = get_all_test(file_path, i[0])
            case_data = {'plan_name': plan_name, 'project_phase': project_phase, 'project_name': project_name,
                         'sheet_name': i[0], 'tester': userid,
                         'workloading': i[2], 'cases': all_case, 'model_name': model_names, 'filename': file_path}
            http_manager.post_data('/insert_case', data=case_data, token=token)
    elif template == 'power':
        logger.info("电源模版")
        raise logger.error("根据要求，必须使用 tdms 导出的测试计划模版！")
        # data = read_test_cases_from_excel(file_path)
        # for sheet, cases in data.items():
        #     case_data = {'filename': file_path, 'sheet_name': sheet, 'project_name': 'power-project',
        #                  'tester': userid, 'workloading': '100(Min)', 'cases': cases}
        #     http_manager.post_data('/insert_case_by_power', data=case_data, token=token)
    else:
        logger.error(template)
        raise template


if __name__ == '__main__':
    file_name = r'C:\Users\71958\Downloads\TDMS\TDMS\Mouse Basic function test 10PB长时间运行.xlsx'
    run_main(file_name, 'ysq', 'aaa')
