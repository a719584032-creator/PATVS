import sys
import openpyxl
from datetime import datetime, timedelta
import random
from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton, QLineEdit, QFileDialog, QVBoxLayout, QHBoxLayout, QMessageBox
)

def random_time_between(start_time_str, end_time_str):
    start = datetime.strptime(start_time_str, '%H:%M:%S')
    end = datetime.strptime(end_time_str, '%H:%M:%S')
    rand_seconds = random.randint(0, int((end - start).total_seconds()))
    return (start + timedelta(seconds=rand_seconds)).strftime('%H:%M:%S')

def process_file(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    data = rows[7:]

    header = data[0]
    idx_name = header.index('职工姓名')
    idx_date = header.index('进入日期')
    idx_time = header.index('进入时间')
    idx_inout = header.index('进入/退出')
    idx_happened = header.index('发生于')

    records = []
    for row in data[1:]:
        if row[idx_name] is None:
            continue
        records.append({
            '姓名': row[idx_name],
            '日期': row[idx_date],
            '时间': row[idx_time],
            '进出': row[idx_inout],
            '发生于': row[idx_happened],
        })

    # 分组
    result_dict = {}
    for rec in records:
        key = (rec['姓名'], rec['日期'])
        if key not in result_dict:
            result_dict[key] = []
        result_dict[key].append(rec)

    result_rows = []
    for (name, date), recs in result_dict.items():
        # 所有in和out
        in_list = [r for r in recs if r['进出'] == 'in']
        out_list = [r for r in recs if r['进出'] == 'out']

        # 第一次in
        if in_list:
            first_in = sorted(in_list, key=lambda x: x['时间'])[0]
        else:
            first_in = {'时间': '', '发生于': ''}
        # 最后一次out
        if out_list:
            last_out = sorted(out_list, key=lambda x: x['时间'])[-1]
        else:
            last_out = {'时间': '', '发生于': ''}
        # 最后一次刷卡
        last_swipe = sorted(recs, key=lambda x: x['时间'])[-1] if recs else {'时间': ''}


        in_time_str = str(first_in['时间'])
        out_time_str = str(last_out['时间'])
        last_swipe_time = str(last_swipe['时间'])

        # 计算工时
        try:
            in_dt = datetime.strptime(in_time_str, '%H:%M:%S')
            out_dt = datetime.strptime(out_time_str, '%H:%M:%S')
            hours = (out_dt - in_dt).total_seconds() / 3600
            if hours < 0:
                hours += 24
            hours = round(hours, 2)
        except Exception:
            hours = ''

        # 计算最后一次刷卡与第一次in的时间差
        try:
            swipe_dt = datetime.strptime(last_swipe_time, '%H:%M:%S')
            in_dt = datetime.strptime(in_time_str, '%H:%M:%S')
            swipe_diff = (swipe_dt - in_dt).total_seconds() / 3600
            if swipe_diff < 0:
                swipe_diff += 24
            swipe_diff = round(swipe_diff, 2)
        except Exception:
            swipe_diff = ''

        result_rows.append([
            name,
            date,
            in_time_str,
            first_in['发生于'],
            out_time_str,
            last_out['发生于'],
            last_swipe_time,
            hours,
            swipe_diff
        ])

    # 写入新Excel
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.append([
        '姓名', '日期', '第一次in时间', '第一次in发生于', '最后一次out时间', '最后一次out发生于',
        '最后一次刷卡时间', '工时(小时)', '最后一次刷卡与第一次in时间差(小时)'
    ])
    for row in result_rows:
        ws_out.append(row)
    wb_out.save(output_path)


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("员工进出统计工具")
        self.resize(580, 280)

        # 输入文件
        self.input_label = QLabel("原始Excel文件：")
        self.input_edit = QLineEdit()
        self.input_btn = QPushButton("选择文件")
        self.input_btn.clicked.connect(self.choose_input)

        # 输出文件
        self.output_label = QLabel("导出到Excel文件：")
        self.output_edit = QLineEdit()
        self.output_btn = QPushButton("选择文件")
        self.output_btn.clicked.connect(self.choose_output)

        # 处理按钮
        self.run_btn = QPushButton("开始统计并导出")
        self.run_btn.clicked.connect(self.run_process)

        # 布局
        input_layout = QHBoxLayout()
        input_layout.addWidget(self.input_edit)
        input_layout.addWidget(self.input_btn)

        output_layout = QHBoxLayout()
        output_layout.addWidget(self.output_edit)
        output_layout.addWidget(self.output_btn)

        main_layout = QVBoxLayout()
        main_layout.addWidget(self.input_label)
        main_layout.addLayout(input_layout)
        main_layout.addWidget(self.output_label)
        main_layout.addLayout(output_layout)
        main_layout.addWidget(self.run_btn)

        self.setLayout(main_layout)

    def choose_input(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择原始Excel文件", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            self.input_edit.setText(file_path)

    def choose_output(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "选择导出Excel文件", "", "Excel Files (*.xlsx)")
        if file_path:
            self.output_edit.setText(file_path)

    def run_process(self):
        input_path = self.input_edit.text().strip()
        output_path = self.output_edit.text().strip()
        if not input_path or not output_path:
            QMessageBox.warning(self, "提示", "请先选择输入和输出文件！")
            return
        try:
            process_file(input_path, output_path)
            QMessageBox.information(self, "完成", "统计并导出成功！")
        except Exception as e:
            QMessageBox.critical(self, "出错", f"发生错误：{e}")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())
