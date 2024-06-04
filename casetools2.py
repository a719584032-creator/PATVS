from openpyxl import load_workbook
from openpyxl import Workbook
from common.rw_excel import MyExcel
import mysql.connector
from mysql.connector import errorcode

# 检查合并单元格是否符合特定规则
def is_merged_within_range(merged_ranges, row, start_col, end_col):
    for rng in merged_ranges:
        # 检查传入的行号是否在某个合并单元格的范围内，同时列号也要在指定的合并列范围内。
        if row >= rng.min_row and row <= rng.max_row and start_col >= rng.min_col and end_col <= rng.max_col:
            return True
    return False


def get_all_test(file_name, sheet_name):
    """
    获取所有用例，按照指定格式[A-F合并为用例标题，A-D列合并为用例步骤，D-F合并为预期结果]
    """
    workbook = load_workbook(file_name)
    sheet = workbook[sheet_name]
    # 获取工作表中所有合并单元格的范围信息，并存入merged_ranges列表。
    merged_ranges = list(sheet.merged_cells.ranges)
    print(merged_ranges)
    # 获取第19行机型数据
    row_19_data = [cell.value for cell in sheet[19] if cell.value is not None]
    model_name = row_19_data[2:]

    # 初始化用例信息存储结构
    case = {
        'title': None,
        'steps': None,
        'expected': None
    }
    all_case = []
    current_title = None
    # 遍历工作表中的所有行
    for row in range(20, sheet.max_row + 1):

        # 检查是否是标题行：A-F列中只有此行数据
        if is_merged_within_range(merged_ranges, row, 1, 6):
            title_cell = sheet.cell(row, 1)
            print("-" * 40)
            print(f"用例标题: {title_cell.value}")
            current_title = title_cell.value
            continue  # 跳过后续步骤，继续下一行

        # 检查是否是步骤：A-D列合并
        if is_merged_within_range(merged_ranges, row, 1, 4):
            steps_cell = sheet.cell(row, 1)
            print(f"用例步骤: {steps_cell.value}")
            case['steps'] = steps_cell.value
        # 检查是否是预期结果：E-F列合并
        if is_merged_within_range(merged_ranges, row, 5, 6):
            expected_cell = sheet.cell(row, 5)
            print(f"预期结果: {expected_cell.value}")
            case['expected'] = expected_cell.value
            print("-" * 40)  # 打印分隔线，表示步骤和预期结果的结束
            # 过滤掉都为 None 的数据，添加后重置字典
            if any(case.values()):
                filled_case =  {
                    'title': current_title if current_title is not None else 'NA',
                    'steps': case['steps'] if case['steps'] is not None else 'NA',
                    'expected': case['expected'] if case['expected'] is not None else 'NA'
                }
                print('************************************************')
                print(filled_case['expected'])
                print('************************************************')
                all_case.append(filled_case)
                case = {
                    'title': None,
                    'steps': None,
                    'expected': None
                }
    return model_name, all_case


def save_cases_to_excel(cases, filename, tester, model_name=None):
    """
    写入新的用例格式，按照机型循环写入用例，sheet页面为 tester
    """
    # 创建一个新工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = tester

    # 添加列标题
    ws.append(['测试机型', '用例标题', '用例步骤', '预期结果'])
    if model_name:
        for m in model_name:
            # 添加用例数据
            for case in cases:
                ws.append([m, case['title'], case['steps'], case['expected']])
    else:
        for case in cases:
            ws.append(['', case['title'], case['steps'], case['expected']])

    # 保存工作簿
    wb.save(filename + '.xlsx')

def insert_data_to_db(plan_name, project_name, sheet_name, tester, workloading, filename, cases, model_name):
    try:
        # 连接到数据库
        connection = mysql.connector.connect(
            host="rm-cn-lf63r60vh0003gto.rwlb.rds.aliyuncs.com",
            user="yesq3_lenovo",
            password="patvs_Lenovo",
            database="lenovoDb"
        )
        cursor = connection.cursor()

        # 查询是否已经存在相同的 plan_name
        cursor.execute("SELECT id FROM TestPlan WHERE plan_name = %s", (plan_name,))
        result = cursor.fetchone()

        if result:
            # 已经存在相同的 plan_name，获取其 id
            plan_id = result[0]
        else:
            # 插入新的 TestPlan 记录
            plan_query = """
                INSERT INTO TestPlan (plan_name, filename)
                VALUES (%s, %s)
            """
            cursor.execute(plan_query, (plan_name, filename))
            plan_id = cursor.lastrowid

        # 检查是否已经存在相同的 sheet_name
        cursor.execute("SELECT id FROM TestSheet WHERE sheet_name = %s AND plan_id = %s", (sheet_name, plan_id))
        sheet_result = cursor.fetchone()

        if sheet_result:
            print(f"Sheet '{sheet_name}' already exists for plan '{plan_name}', skipping insertion of sheet and cases.")
            return  # 如果 sheet_name 已经存在，直接跳过后续操作
        else:
            # 插入新的 TestSheet 记录
            sheet_query = """
                INSERT INTO TestSheet (sheet_name, project_name, tester, workloading, plan_id)
                VALUES (%s, %s, %s, %s, %s)
            """
            cursor.execute(sheet_query, (sheet_name, project_name, tester, workloading, plan_id))
            sheet_id = cursor.lastrowid

            # 插入到 TestCase 表
            case_query = """
                INSERT INTO TestCase (ModelName, CaseTitle, CaseSteps, ExpectedResult, sheet_id)
                VALUES (%s, %s, %s, %s, %s)
            """
            for case in cases:
                for model in model_name:
                    cursor.execute(case_query, (model, case['title'], case['steps'], case['expected'], sheet_id))

        # 提交事务
        connection.commit()
    except mysql.connector.Error as err:
        print(f"Error: {err}")
        connection.rollback()
    finally:
        cursor.close()
        connection.close()
def validate_excel_format(file_name):
    """
    校验 Excel 文件格式
    """
    workbook = load_workbook(file_name)
    required_sheets = ['Plan Information', 'Case List']

    for sheet_name in required_sheets:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"缺少必要的 sheet 页: {sheet_name}")

    plan_info_sheet = workbook['Plan Information']
    case_list_sheet = workbook['Case List']

    if plan_info_sheet.cell(1, 1).value != 'Plan name':
        raise ValueError("Plan Information sheet 页中 A1 单元格的值必须为 'Plan name'")

    if case_list_sheet.cell(2, 2).value != 'Case name':
        raise ValueError("Case List sheet 页中 B2 单元格的值必须为 'Case name'")

    print("文件格式校验通过")

def run_main(file_name):
#    validate_excel_format(file_name)
    data = MyExcel(file_name)
    data.active_sheet('Plan Information')
    plan_name = data.get_value_by_rc(1, 2)
    print(plan_name)
    project_name = data.get_value_by_rc(1, 4)
    print(project_name)
    data.active_sheet('Case List')
    all_sheet = data.getColValues(2)[2:]
    # 实际 sheet_name 需要加上递增的前缀
    all_sheet_with_prefix = [f'{i + 1}-{val}' for i, val in enumerate(all_sheet)]
    all_tester = data.getColValues(14)[2:]
    all_workloading = data.getColValues(15)[2:]
    print(all_sheet_with_prefix)
    print(all_tester)
    sheet_and_tester_and_workloading = list(zip(all_sheet_with_prefix, all_tester, all_workloading))
    print(sheet_and_tester_and_workloading)
    for i in sheet_and_tester_and_workloading:
        model_name, all_case = get_all_test(file_name, i[0])
    #    save_cases_to_excel(all_case, i[0], i[1], model_name)
        insert_data_to_db(plan_name, project_name, i[0], i[1], i[2], file_name, all_case, model_name)



if __name__ == '__main__':
    file_name = 'TestPlanWithResult_M410_Mouse_PATVS软件测试(1)_20240528100806.xlsx'
    run_main(file_name)


    # data = MyExcel(file_name)
    # data.active_sheet('Case List')
    # sheet_name = '2-KB test  basic function'
    # model_name, all_case = get_all_test(file_name, sheet_name)
    # model = ['qqq', 'mmm']

    # save_cases_to_excel(all_case, sheet_name, 'ysq', model)
