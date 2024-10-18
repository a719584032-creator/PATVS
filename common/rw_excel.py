# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
# from common.public.tools import Public
from common.logs import logger
import re


class MyExcel:
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        self.wb.guess_types = True

    def active_sheet(self, sheet_name):
        """
        切换到指定名字的sheet页面
        :param sheet_name:
        :return:
        """
        self.ws = self.wb[sheet_name]

    def get_sheet_names(self):
        """
         获取所有的sheet name
        :return:
        """
        sheet_names = self.wb.sheetnames
        return sheet_names

    def get_value(self, add):
        return self.ws[add].value

    def get_value_by_rc(self, r, c):
        """
        获取指定单元格值
        :param r:
        :param c:
        :return:
        """
        return self.ws.cell(r, c).value

    def set_value(self, add, value):
        if 'H' in add:
            self.ws[add].alignment = Alignment(horizontal='center', vertical='center')
            self.ws[add].border = Border(bottom=Side(style='thin'))
            self.ws[add].font = Font(size=10)
        self.ws[add] = value

    def set_value_by_rc(self, row, column, value):
        """
        往指定单元格写入值
        :param row:
        :param column:
        :param value:
        :return:
        """
        self.ws.cell(row=row, column=column).value = value

    def getRowValues(self, row):
        """
        获取某行所有值
        :param row:
        :return:
        """
        columns = self.ws.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    def get_appoint_row_values(self, row_start, row_end=None):
        """
        获取指定行 - 指定行的数据
        :param row_start: 起始行
        :param row_end: 结束行
        :return:
        """
        if row_end is None:
            row_end = self.get_max_row()
        data = []
        for i in range(row_start, row_end + 1):
            data.append(self.getRowValues(i))
        return data

    def get_abc(self, r, c):
        """
        把字符串 'a,b,c' 切割成列表 ['a','b','c'] 返回
        :param r:
        :param c:
        :return:
        """
        value = self.get_value_by_rc(r, c)
        result = re.split(r'[，.,\s]+', value.lower())
        result_data = [x for x in result if x != '']
        return result_data

    def getColValues(self, col):
        """
        获取某列所有值
        :param col:
        :return:
        """
        rowumns = self.ws.max_row
        coldata = []
        for i in range(1, rowumns + 1):
            cellvalue = self.ws.cell(row=i, column=col).value
            coldata.append(cellvalue)
        return coldata

    def save_file(self, add_name):
        self.wb.save(add_name)

    def remove_sheet(self, sheetname):
        """
        删除sheet
        """
        ws = self.wb[sheetname]
        self.wb.remove(ws)

    def do_close(self):
        self.wb.close()

    def create_sheet(self, sheet_name):
        self.ws = self.wb.create_sheet(sheet_name)
        return self.ws

    def copy_sheet(self, worksheet):
        self.wb.copy_worksheet(worksheet)

    def get_max_row(self):
        return self.ws.max_row

    def get_max_colum(self):
        return self.ws.max_column

    def validate_case_data(self, col_data):
        # 格式检验
        expected_cols = ['测试机型', '用例标题', '用例步骤', '预期结果']
        logger.info(col_data)
        if col_data != expected_cols:
            raise ValueError("列标题的格式错误！")
        return True


if __name__ == '__main__':
    path = Public.get_root_path() + '/data/template for packaging collection- thinkplus TF10.xlsx'
    data = MyExcel(path)
    data.active_sheet('1200x1000mm栈板')
    num = data.get_max_row()
    sales_start_index = 0
    sales_end_index = 0
    transport_start_index = 0
    transport_end_index = 0
    for i in range(18, num):
        value = data.getRowValues(i)
        if i == 19:
            print(value)
        if 'sales packaging' in value:
            sales_start_index = i + 2
            logger.info(f" sales packaging 起始数据行 {sales_start_index}")
            continue
        if value[1] is None or value[1] == '':
            sales_end_index = i - 1
            logger.info(f" sales packaging 结束数据行 {sales_end_index}")
            break
    else:
        logger.error("sales packaging 数据查找异常")
    flg = 0
    for i in range(18, num):
        value = data.getRowValues(i)
        if 'transport packaging' in value:
            flg = i
            logger.info(f"transport 关键词已查到,在 {flg} 行")
            break
    else:
        logger.error("transport packaging 数据查找异常")
    for i in range(flg, num):
        value = data.getRowValues(i)
        if 'transport packaging' in value:
            transport_start_index = i + 2
            logger.info(f" transport packaging 起始数据行 {transport_start_index}")
            continue
        if value[1] is None or value[1] == '':
            transport_end_index = i - 1
            logger.info(f" transport packaging 结束数据行 {transport_end_index}")
            break
    else:
        logger.error("sales packaging 数据查找异常")
    sales_packaging_data = data.get_appoint_row_values(sales_start_index, sales_end_index)
    transport_packaging_data = data.get_appoint_row_values(transport_start_index, transport_end_index)
    print(sales_packaging_data)
    print(len(sales_packaging_data))
    print(transport_packaging_data)
    print(len(transport_packaging_data))
    for item in transport_packaging_data:
        if item[1] == 'Pallet':
            item[1] = 'Palette'
        if item[2] == 'Plastic - PP(polypropylene)':
            item[2] = 'Plastic - PP (polypropylene)'
    print(transport_packaging_data)
    print('sales packaging' == 'sales packaging')
