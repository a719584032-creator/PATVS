import io
import json
import sys
import os
import zipfile

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import boto3
from botocore.client import Config
import jwt
import datetime
import re
import uuid
import time
from flask import Flask, request, jsonify, send_file
from common.logs import logger
from config_manager.config import env_config
from mysql.connector.pooling import MySQLConnectionPool
from patvsweb_services.sql_manager import TestCaseManager
from functools import wraps
from werkzeug.middleware.proxy_fix import ProxyFix
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
app.config['SECRET_KEY'] = 'lenovo_secret_key'



# 配置S3存储桶
BUCKET_NAME = env_config.global_setting.aws_bucket_name

# 配置S3客户端
s3 = boto3.client('s3',
                  aws_access_key_id=env_config.global_setting.aws_access_key,
                  aws_secret_access_key=env_config.global_setting.aws_secret_key,
                  endpoint_url=env_config.global_setting.aws_endpoint_url,
                  region_name=env_config.global_setting.aws_region_name,
                  config=Config(signature_version=env_config.global_setting.aws_signature_version))

# 数据库配置
DB_CONFIG = {
    'host': env_config.global_setting.db_host,
    'user': env_config.global_setting.db_user,
    'password': env_config.global_setting.db_password,
    'database': env_config.global_setting.db_database,
    'buffered': env_config.global_setting.db_buffered
}

# 生产
# DB_CONFIG = {
#     'host': '10.196.155.148',
#     'user': 'a_appconnect',
#     'password': 'dHt6BGB4Zxi^',
#     'database': 'patvs_db',
#     'buffered': True
# }

db_pool = MySQLConnectionPool(pool_name="mypool", pool_size=10, **DB_CONFIG)


def get_db_connection():
    return db_pool.get_connection()


def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('x-access-tokens')
        if not token:
            return jsonify({'error': 'Token is missing!'}), 403
        try:
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=["HS256"])
            current_user = data['username']
            current_userid = data['userid']
            logger.info(f"User: {current_user}, UserID: {current_userid}")
        except:
            return jsonify({'error': 'Token is invalid!'}), 403
        return f(*args, current_user=current_user, current_userid=current_userid, **kwargs)

    return decorated


@app.route('/update_start_time', methods=['POST'])
@token_required
def update_start_time(current_user, current_userid):
    data = request.json
    case_id = data.get('case_id')
    model_id = data.get('model_id')
    start_time = data.get('start_time')
    logger.info(f"Updating start time for case_id: {case_id}, model_id: {model_id}")
    if not case_id or not model_id:
        return jsonify({'error': 'Missing required parameters'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        now = datetime.datetime.now()
        start_time = now.strftime('%Y-%m-%d %H:%M:%S')
        manager.update_start_time_by_case_id(case_id, model_id, start_time)
        conn.commit()
        return jsonify({'message': 'Start time updated successfully.'}), 200
    except Exception as e:
        conn.rollback()
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/update_end_time', methods=['POST'])
@token_required
def update_end_time(current_user, current_userid):
    data = request.json
    case_id = data.get('case_id')
    model_id = data.get('model_id')
    case_result = data.get('case_result')
    comment = data.get('comment', None)
    logger.info(
        f"Updating end time for case_id: {case_id}, model_id: {model_id}, actions: {case_result}, comment: {comment}")
    if not case_id or not case_result:
        return jsonify({'error': 'Missing required parameters'}), 400
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        manager.update_end_time_case_id(case_id, model_id, case_result, comment)
        conn.commit()
        return jsonify({'message': 'End time updated successfully.'}), 200
    except Exception as e:
        conn.rollback()
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/insert_case', methods=['POST'])
@token_required
def insert_case(current_user, current_userid):
    data = request.json
    plan_name = data.get('plan_name')
    project_name = data.get('project_name')
    project_phase = data.get('project_phase')
    sheet_name = data.get('sheet_name')
    userid = data.get('tester')
    workloading = data.get('workloading')
    filename = data.get('filename')
    cases = data.get('cases')
    model_name = data.get('model_name')

    logger.info(f"Inserting case for plan: {plan_name}, project: {project_name}, sheet: {sheet_name}")

    if not plan_name or not project_name or not project_phase or not sheet_name or not userid or not workloading or not filename or not cases or not model_name:
        return jsonify({'error': 'Missing required parameters'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        manager.insert_case_by_filename(plan_name, project_name, project_phase, sheet_name, userid, workloading,
                                        filename, cases,
                                        model_name)
        conn.commit()
        return jsonify({'message': 'Case inserted successfully.'}), 200
    except Exception as e:
        conn.rollback()
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/insert_case_by_power', methods=['POST'])
@token_required
def insert_case_by_power(current_user, current_userid):
    data = request.json
    project_name = data.get('project_name')
    sheet_name = data.get('sheet_name')
    tester = data.get('tester')
    workloading = data.get('workloading')
    filename = data.get('filename')
    cases = data.get('cases')

    logger.info(f"Inserting case for file: {filename}, project: {project_name}, sheet: {sheet_name}")

    if not project_name or not sheet_name or not tester or not workloading or not filename or not cases:
        return jsonify({'error': 'Missing required parameters'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        manager.insert_case_by_power_filename(filename, sheet_name, project_name, tester, workloading, cases)
        conn.commit()
        return jsonify({'message': 'Case inserted successfully.'}), 200
    except Exception as e:
        conn.rollback()
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_comments', methods=['POST'])
def get_comments_for_case():
    data = request.json
    execution_ids = data.get('execution_ids')
    if not execution_ids:
        return jsonify({'error': 'Missing required parameters'}), 400
    logger.info(f"Fetching cases for execution_ids: {execution_ids}")
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        comments = manager.select_all_comments(execution_ids)
        logger.info(comments)
        return jsonify({'comments': comments}), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_project_names/<int:userid>', methods=['GET'])
def get_project_names(userid):
    logger.info(f"Fetching project names for userid: {userid}")
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        project_names = manager.select_all_project_names_by_username(userid)
        return jsonify({'project_names': project_names}), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_plan_names/<int:userid>/<string:project_name>', methods=['GET'])
def get_plan_names(userid, project_name):
    logger.info(f"Fetching plan project name for : {project_name}, usrid for : {userid}")
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        plan_names = manager.select_all_plan_names_by_project(userid, project_name)
        return jsonify({'plan_names': plan_names}), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_model_names/<int:plan_id>', methods=['GET'])
def get_model_names(plan_id):
    logger.info(f"Fetching plan id for : {plan_id}")
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        model_names = manager.select_all_model_names_by_plan_id(plan_id)
        return jsonify({'model_names': model_names}), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_sheet_names/<int:plan_id>', methods=['GET'])
def get_sheet_names(plan_id):
    logger.info(f"Fetching plan id for : {plan_id}")
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        sheet_names_with_ids = manager.select_all_sheet_names_by_plan_id(plan_id)
        return jsonify({'sheet_names_with_ids': sheet_names_with_ids}), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_plan_names_by_admin', methods=['GET'])
def get_plan_names_by_admin():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        plan_names = manager.select_all_plan_names()
        return jsonify({'plan_names': plan_names}), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_sheet_names_by_admin/<string:plan_name>', methods=['GET'])
def get_sheet_names_by_admin(plan_name):
    logger.info(f"Fetching sheet names for plan_name: {plan_name}")
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        sheet_names_with_ids = manager.select_all_sheet_names_by_plan(plan_name)
        return jsonify({'sheet_names_with_ids': sheet_names_with_ids}), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_userid/<string:username>', methods=['GET'])
def get_userid(username):
    logger.info(f"Fetching user ID for username: {username}")
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        user_id = manager.select_userid_by_username(username)
        return jsonify({'user_id': user_id}), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_filename/<string:filename>', methods=['GET'])
def get_filename(filename):
    logger.info(f"Fetching filename: {filename}")
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        file_exists = manager.select_filename_by_filename(filename)
        return jsonify({'file_exists': file_exists}), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_plan_name_by_planname/<string:plan_name>/<int:user_id>', methods=['GET'])
def get_plan_name_by_planname(plan_name, user_id):
    logger.info(f"Fetching plan_name: {plan_name}, userid: {user_id}")
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        plan_exists = manager.select_plan_name_by_plan_name(plan_name, user_id)
        return jsonify({'plan_exists': plan_exists}), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_plan_name/<string:filename>', methods=['GET'])
def get_plan_name(filename):
    logger.info(f"Fetching plan name for filename: {filename}")
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        plan_name = manager.select_plan_name_by_filename(filename)
        return jsonify({'plan_name': plan_name}), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_cases_status/<int:sheet_id>/<int:model_id>', methods=['GET'])
def get_cases_status(sheet_id, model_id):
    logger.info(f"Fetching cases for sheet_id: {sheet_id}, model_id: {model_id}")
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        cases = manager.select_case_status(model_id, sheet_id)
        formatted_cases = []
        for case in cases:
            comment = case[4]
            # 检查 comment 是否为 "N/A: No Comment"，如果是则替换为 None
            if comment == "N/A: No Comment":
                comment = None
            case_dict = {
                'TestResult': case[0],
                'TestTime': case[1],
                'StartTime': case[2].strftime('%Y-%m-%d %H:%M:%S') if case[2] else None,
                'EndTime': case[3].strftime('%Y-%m-%d %H:%M:%S') if case[3] else None,
                'Comment': comment,
                'CaseTitle': case[5],
                'PreConditions': case[6],
                'CaseSteps': case[7],
                'ExpectedResult': case[8],
                'ExecutionID': case[9],
                'ModelID': case[10],
                'CaseID': case[11],
                'SheetID': case[12],
                'FailCount': case[13],
                'BlockConut': case[14]
            }
            formatted_cases.append(case_dict)
        return jsonify(formatted_cases), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/export_plan/<int:plan_id>', methods=['GET'])
def export_plan(plan_id):
    logger.info(f"Fetching cases for plan_id: {plan_id}")
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        plan_name, plan_data = manager.select_case_status_by_plan_id(plan_id)
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for model_name, sheets in plan_data.items():
                logger.warning(f"model_name: {model_name}, sheet_names: {list(sheets.keys())}")
                wb = Workbook()
                default_sheet = wb.active
                created_sheet = False
                for sheet_name, cases in sheets.items():
                    logger.warning(f"  sheet: {sheet_name}, case count: {len(cases)}")
                    ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet名最长31字符
                    ws.append([
                        '测试结果', '测试耗时(S)', '用例标题', '前置条件', '用例步骤', '预期结果',
                        '开始时间', '完成时间', '评论', '失败次数', '阻塞次数', '查看图片'
                    ])
                    # 1. 批量收集execution_id
                    execution_id_list = []
                    for case in cases:
                        execution_id = case[9]  # 最后一个字段
                        if execution_id:
                            execution_id_list.append(execution_id)

                    # 2. 一次性查出所有图片信息
                    images_map = {}
                    if execution_id_list:
                        images_map = manager.select_images_by_execution_ids(execution_id_list, request.host_url)

                    # 3. 逐行写入
                    for case in cases:
                        (
                            test_result, test_time, start_time, end_time, comment,
                            case_title, pre_conditions, case_steps, expected_result,
                            execution_id, model_id, case_id, sheet_id, fail_count, block_count
                        ) = case
                        # 格式化时间
                        start_time_fmt = start_time.strftime('%Y-%m-%d %H:%M:%S') if start_time else ''
                        end_time_fmt = end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else ''
                        # Comment处理
                        comment_fmt = '' if comment == "N/A: No Comment" else comment
                        # 取图片
                        image_data = ''
                        if execution_id and images_map:
                            images = images_map.get(str(execution_id), [])
                            if images:
                                image_data = json.dumps(images, ensure_ascii=False)  # 保证中文不转义

                        # 写入Excel时不写入ID
                        row_data = [
                            test_result, test_time, case_title, pre_conditions, case_steps, expected_result,
                            start_time_fmt, end_time_fmt, comment_fmt, fail_count, block_count, image_data
                        ]
                        ws.append(row_data)

                        # 设置单元格背景色（注意：ws.append后，当前行号=ws.max_row）
                        fill = None
                        if row_data[0] == 'Pass':
                            fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        elif row_data[0] == 'Fail' or row_data[0] == 'Block':
                            fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")
                        if fill:
                            for cell in ws[ws.max_row]:
                                cell.fill = fill
                    created_sheet = True
                if created_sheet:
                    wb.remove(default_sheet)
                else:
                    default_sheet.title = "无数据"
                    default_sheet.append(["没有可导出的用例"])
                    # 保存到内存
                xlsx_buffer = io.BytesIO()
                wb.save(xlsx_buffer)
                xlsx_buffer.seek(0)
                zip_path = f"{model_name}/{plan_name}.xlsx"
                zip_file.writestr(zip_path, xlsx_buffer.read())
        zip_buffer.seek(0)
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f"{plan_name}.zip"
        )
    except Exception as e:
        logger.error(f"Export error: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_case_result', methods=['GET'])
def get_case_result():
    """
    接口：根据 model_id 和 case_id 或 execution_id 获取用例结果。
    """
    # 获取请求参数
    execution_id = request.args.get('execution_id', default=None, type=int)
    model_id = request.args.get('model_id', default=None, type=int)
    case_id = request.args.get('case_id', default=None, type=int)

    # 日志记录
    logger.info(
        f"Fetching case result with parameters: execution_id={execution_id}, model_id={model_id}, case_id={case_id}")

    # 数据库连接
    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        manager = TestCaseManager(conn, cursor)

        # 根据 execution_id 查询
        if execution_id is not None:
            logger.info(f"Fetching case result for execution_id: {execution_id}")
            case_result = manager.select_case_result_by_execution_id(execution_id)
        # 根据 model_id 和 case_id 查询
        elif model_id is not None and case_id is not None:
            logger.info(f"Fetching case result for model_id: {model_id}, case_id: {case_id}")
            case_result = manager.select_case_result_by_id(case_id, model_id)
        else:
            # 参数不足，返回 400 错误
            logger.warning("Invalid parameters: either execution_id or both model_id and case_id must be provided.")
            return jsonify(
                {'error': 'Invalid parameters. Provide either execution_id or both model_id and case_id.'}), 400

        # 返回查询结果
        return jsonify({'case_result': case_result}), 200

    except Exception as e:
        # 捕获异常并记录错误日志
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500

    finally:
        # 关闭数据库连接
        cursor.close()
        conn.close()


@app.route('/reset_case_result', methods=['POST'])
@token_required
def reset_case_result(current_user, current_userid):
    data = request.json
    execution_id = data.get('execution_id')
    if not execution_id:
        return jsonify({'error': 'Missing required parameters'}), 400
    logger.info(f"Fetching cases for execution_id: {execution_id}")
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        manager.reset_case_by_execution_id(execution_id)
        conn.commit()
        return jsonify({'message': 'Test case reset successfully.'}), 200
    except Exception as e:
        conn.rollback()
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/calculate_progress_and_pass_rate', methods=['GET'])
def calculate_progress_and_pass_rate():
    plan_id = request.args.get('planId')
    model_id = request.args.get('modelId')
    sheet_id = request.args.get('sheetId')
    if not plan_id or not model_id or not sheet_id:
        return jsonify({'error': 'Missing required parameters'}), 400
    logger.info(f"Fetching plan for : {plan_id}")
    logger.info(f"Fetching model for : {model_id}")
    logger.info(f"Fetching sheets for : {sheet_id}")
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        result = manager.calculate_progress_and_pass_rate(plan_id, model_id, sheet_id)
        return jsonify({'result': result}), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/calculate_plan_statistics/<int:plan_id>', methods=['GET'])
def calculate_plan_statistics(plan_id):
    logger.info(f"Fetching plan for plan_id: {plan_id}")
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        result = manager.calculate_plan_statistics(plan_id)
        return jsonify({'result': result}), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_start_time/<int:model_id>/<int:case_id>', methods=['GET'])
def get_start_time(model_id, case_id):
    logger.info(f"Fetching start_time for model_id: {model_id}, case_id: {case_id}")
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        logger.warning(f"传入 model_id: {model_id}, case_id: {case_id}")
        start_time = manager.select_start_time(model_id, case_id)
        if start_time:
            # Convert datetime object to string
            start_time = start_time.strftime('%Y-%m-%d %H:%M:%S')
        logger.warning(start_time)
        return jsonify({'start_time': start_time}), 200
    except Exception as e:
        conn.rollback()
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/add_user', methods=['POST'])
def add_user():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    role = data.get('role', None)

    logger.info(f"add user by username: {username}")

    # 用户名和密码必填校验
    if not username or not password:
        logger.warning('用户名和密码不能为空')
        return jsonify({'error': '用户名和密码不能为空'}), 400

    # 用户名长度校验
    if not (3 <= len(username) <= 15):
        logger.warning('用户名长度必须为3-15位')
        return jsonify({'error': '用户名长度必须为3-15位'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        # 检查用户是否已存在
        if manager.user_exists(username):
            logger.warning('用户名已存在')
            return jsonify({'error': '用户名已存在'}), 400

        manager.add_user(username, password, role)
        conn.commit()
        return jsonify({'message': f'用户 {username} 添加成功！'}), 200
    except Exception as e:
        conn.rollback()
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/login', methods=['POST'])
def validate_user():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    logger.info(f"login username: {username}, password: {password} ")

    if not username or not password:
        return jsonify({'error': 'Missing required parameters'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        valid, role, userid = manager.validate_user(username, password)
        if valid:
            token = jwt.encode({
                'username': username,
                'userid': userid,
                'exp': datetime.datetime.utcnow() + datetime.timedelta(hours=96)
            }, app.config['SECRET_KEY'], algorithm="HS256")
            return jsonify({'token': token, 'role': role, 'userid': userid, 'username': username})
        else:
            return jsonify({'error': 'Invalid credentials'}), 401
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/change_user_password', methods=['POST'])
def change_user_password():
    data = request.json
    username = data.get('username')
    old_password = data.get('old_password')
    new_password = data.get('new_password')
    logger.info(
        f"change_user_password username: {username}, old_password: {old_password} , new_password: {new_password}")

    if not username or not old_password or not new_password:
        return jsonify({'error': 'Missing required parameters'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        result = manager.change_user_password(username, old_password, new_password)
        return jsonify({'result': result}), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_tester', methods=['GET'])
def get_tester():
    plan_name = request.args.get('plan_name')
    sheet_id = request.args.get('sheet_id')
    logger.info(f"get_tester plan_name: {plan_name}, sheet_id: {sheet_id} ")

    # if not plan_name or not sheet_id:
    #     return jsonify({'error': 'Missing required parameters'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        tester = manager.select_tester_by_plan_or_sheet(plan_name, sheet_id)
        return jsonify({'tester': tester}), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_case_actions_and_num/<int:case_id>', methods=['GET'])
def get_case_actions_and_num(case_id):
    logger.info(f"get_case_actions_and_num case_id: {case_id} ")

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        title = manager.select_case_title(case_id)
        logger.info(title)
        # "[S3+5][S4+6][S5+3]N5(03)BT Pairing multi devices BT多设备配对"
        # 提取[]括号内的内容
        bracket_contents = re.findall(r'\[([^\]]+)\]', title)
        logger.warning(bracket_contents)
        # 提取关键参数
        key_params = []
        for content in bracket_contents:
            matches = re.match(r'(.+?)\+(\d+(\.\d+)?)', content)
            if matches:
                key_params.append((matches.group(1), matches.group(2)))
        logger.warning(key_params)
        return jsonify({'actions_and_num': key_params}), 200
    except Exception as e:
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/update_project_workloading_tester', methods=['POST'])
def update_project_workloading_tester():
    data = request.json
    plan_id = data.get('plan_id')
    workloading = data.get('workloading', None)
    tester = data.get('tester', None)
    logger.info(
        f"update_project_workloading_tester plan_id: {plan_id}")
    if not plan_id:
        return jsonify({'error': 'Missing required parameters'}), 400
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        if tester:
            res = manager.select_userid_by_username(tester)
            if not res:
                return jsonify({'message': f'你输入的用户 {tester} 不存在'}), 400
        manager.update_project_workloading_tester(plan_id, workloading, tester)
        conn.commit()
        return jsonify({'message': 'success'}), 200
    except Exception as e:
        conn.rollback()
        logger.error(f"An error occurred: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/hello', methods=['GET'])
def get_hello():
    return jsonify({'tester': "hello,hello,hello,hello,hello。网络是通的。"}), 200


# 生成唯一文件名
def generate_unique_filename(original_filename):
    extension = original_filename.rsplit('.', 1)[1]
    unique_filename = f"{uuid.uuid4()}.{extension}"
    return unique_filename


# @app.route('/upload-images', methods=['POST'])
# @token_required
# def upload_image(current_user):
#     image_files = request.files.getlist('image_files')
#     case_id = request.form.get('case_id')
#     model_id = request.form.get('model_id')
#     case_result = request.form.get('case_result')
#     comment = request.form.get('comment', None)
#
#     if not case_id or not model_id or not case_result or not image_files:
#         return jsonify({'error': 'case_id, model_id, case_result and image files are required'}), 400
#
#     conn = get_db_connection()
#     cursor = conn.cursor()
#     manager = TestCaseManager(conn, cursor)
#
#     try:
#         images_data = []
#
#         for image_file in image_files:
#             if not hasattr(image_file, 'read'):
#                 logger.error("One of the image_files is not a file-like object")
#                 continue
#             # 保留原文件名
#             original_filename = image_file.filename
#             timestamp = int(time.time())
#             unique_filename = f"{timestamp}_{original_filename}"
#
#             today = datetime.datetime.now().strftime("%Y-%m-%d")
#             s3_key = f"{case_id}/{today}/{unique_filename}"
#
#             mime_type = image_file.mimetype
#             image_file.seek(0, os.SEEK_END)
#             file_size = image_file.tell()
#             image_file.seek(0)
#
#             # 上传文件到S3
#             try:
#                 s3.upload_fileobj(
#                     image_file,
#                     BUCKET_NAME,
#                     s3_key,
#                     ExtraArgs={'ContentType': mime_type}
#                 )
#                 logger.info(f"Uploaded file: {original_filename} to S3 with key: {s3_key}")
#             except Exception as e:
#                 logger.error(f"Failed to upload to S3: {e}")
#                 return jsonify({'error': 'Failed to upload to S3'}), 500
#
#             # 准备图片数据
#             image_data = {
#                 'original_file_name': original_filename,
#                 'stored_file_name': unique_filename,
#                 'file_path': s3_key,
#                 'file_size': file_size,
#                 'mime_type': mime_type
#             }
#             images_data.append(image_data)
#
#         # 插入执行记录和所有图片信息
#         try:
#             execution_id = manager.insert_execution_with_image(case_id, model_id, case_result, images_data, comment)
#             conn.commit()
#             return jsonify({'execution_id': execution_id, 'uploaded_images': images_data}), 200
#         except Exception as e:
#             logger.error(f"Failed to insert execution and image records: {e}")
#             conn.rollback()
#             return jsonify({'error': 'Failed to insert execution and image records'}), 500
#
#     finally:
#         cursor.close()
#         conn.close()
@app.route('/upload-images', methods=['POST'])
@token_required
def upload_image(current_user, current_userid):
    image_files = request.files.getlist('image_files')
    case_id = request.form.get('case_id')
    model_id = request.form.get('model_id')
    case_result = request.form.get('case_result')
    comment = request.form.get('comment', None)

    if not case_id or not model_id or not case_result or not image_files:
        return jsonify({'error': 'case_id, model_id, case_result and image files are required'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    manager = TestCaseManager(conn, cursor)

    try:
        images_data = []
        # 定义存储文件的根目录为 /data
        upload_root = env_config.global_setting.image_path

        for image_file in image_files:
            if not hasattr(image_file, 'read'):
                logger.error("One of the image_files is not a file-like object")
                continue

            # 保留原文件名
            original_filename = image_file.filename
            timestamp = int(time.time())
            unique_filename = f"{timestamp}_{original_filename}"

            today = datetime.datetime.now().strftime("%Y-%m-%d")
            # 构建文件存储路径
            directory_path = os.path.join(upload_root, today, case_id)
            os.makedirs(directory_path, exist_ok=True)  # 如果目录不存在则创建
            file_path = os.path.join(directory_path, unique_filename)

            mime_type = image_file.mimetype
            image_file.seek(0, os.SEEK_END)
            file_size = image_file.tell()
            image_file.seek(0)

            # 将文件保存到 /data 分区
            try:
                image_file.save(file_path)
                logger.info(f"Saved file: {original_filename} to {file_path}")
            except Exception as e:
                logger.error(f"Failed to save file to server: {e}")
                return jsonify({'error': 'Failed to save file to server'}), 500

            # 准备图片数据
            image_data = {
                'original_file_name': original_filename,
                'stored_file_name': unique_filename,
                'file_path': file_path,
                'file_size': file_size,
                'mime_type': mime_type
            }
            images_data.append(image_data)

        # 插入执行记录和所有图片信息
        try:
            execution_id = manager.insert_execution_with_image(case_id, model_id, case_result, images_data, comment)
            conn.commit()
            return jsonify({'execution_id': execution_id, 'uploaded_images': images_data}), 200
        except Exception as e:
            logger.error(f"Failed to insert execution and image records: {e}")
            conn.rollback()
            return jsonify({'error': 'Failed to insert execution and image records'}), 500

    finally:
        cursor.close()
        conn.close()


@app.route('/get_execution_ids', methods=['POST'])
@token_required
def get_execution_ids(current_user, current_userid):
    data = request.json
    case_ids = data.get('case_ids')
    model_id = data.get('model_id')
    logger.info(f"case_ids is {case_ids}, model_id is {model_id}")
    if not case_ids or not model_id:
        return jsonify({'error': 'case_ids, model_id are required'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    manager = TestCaseManager(conn, cursor)
    try:
        execution_ids = manager.select_execution_ids(case_ids, model_id)
        return jsonify({'execution_ids': execution_ids}), 200
    except Exception as e:
        logger.error(f"Failed to insert execution and image records: {e}")
        return jsonify({'error': 'Failed to select executionids'}), 500


# @app.route('/get_images/<int:execution_id>', methods=['GET'])
# def get_images(execution_id):
#     if not execution_id:
#         return jsonify({'error': 'execution_id is required'}), 400
#
#     conn = get_db_connection()
#     cursor = conn.cursor()
#     manager = TestCaseManager(conn, cursor)
#     try:
#         # 查询数据库获取图片信息
#         images = manager.select_images_by_execution_id(execution_id)
#         if not images:
#             return jsonify({'error': 'No images found for the given execution_id'}), 404
#
#         # 构建返回的图片信息列表，包括预签名 URL
#         images_data = []
#         for image in images:
#             file_path = image[3]
#             presigned_url = generate_presigned_url(file_path)
#
#             if not presigned_url:
#                 logger.error(f"Failed to generate presigned URL for {file_path}")
#                 continue
#
#             image_info = {
#                 'original_file_name': image[2],
#                 'stored_file_name': image[7],
#                 'file_path': file_path,
#                 'file_size': image[4],
#                 'mime_type': image[5],
#                 'time': image[6].strftime('%Y-%m-%d %H:%M:%S') if image[6] else None,
#                 'url': presigned_url  # 添加预签名 URL
#             }
#             images_data.append(image_info)
#
#         logger.info(images_data)
#         return jsonify({'execution_id': execution_id, 'images': images_data}), 200
#
#     except Exception as e:
#         logger.error(f"Failed to retrieve images: {e}")
#         return jsonify({'error': 'Failed to retrieve images'}), 500
#
#     finally:
#         cursor.close()
#         conn.close()
#
#
# def generate_presigned_url(object_key, expiration=604800):  # 默认一周有效
#     try:
#         url = s3.generate_presigned_url('get_object',
#                                         Params={
#                                             'Bucket': BUCKET_NAME,
#                                             'Key': object_key,
#                                             'ResponseContentType': 'image/jpeg',  # 根据实际情况调整 MIME 类型
#                                             'ResponseContentDisposition': 'inline'  # 设置为 inline 以便浏览器预览
#                                         },
#                                         ExpiresIn=expiration)
#         return url
#     except Exception as e:
#         logger.error(f"Error generating presigned URL: {e}")
#         return None
@app.route('/get_images', methods=['POST'])
def get_images():
    data = request.json
    execution_ids = data.get('execution_ids')
    if not execution_ids or not isinstance(execution_ids, list):
        return jsonify({'error': 'execution_ids must be a non-empty list'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    manager = TestCaseManager(conn, cursor)
    try:
        images_data = manager.select_images_by_execution_ids(execution_ids, request.host_url)
        return jsonify({'images': images_data}), 200
    except Exception as e:
        logger.error(f"Failed to retrieve images: {e}")
        return jsonify({'error': 'Failed to retrieve images'}), 500
    finally:
        cursor.close()
        conn.close()


# 提供静态文件服务
@app.route('/uploads/<path:filename>', methods=['GET'])
def serve_file(filename):
    # 基于 /data/uploads 提供文件服务
    upload_root = env_config.global_setting.image_path
    file_path = os.path.join(upload_root, filename)

    if not os.path.exists(file_path):
        return jsonify({'error': 'File not found'}), 404

    # 返回文件内容
    return send_file(file_path)


@app.route('/modify/case_titles', methods=['POST'])
def modify_case_titles():
    data = request.json
    cases = data.get("cases")
    if not cases or not isinstance(cases, list):
        return jsonify({'error': 'cases(list) is required'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()
    manager = TestCaseManager(conn, cursor)
    try:
        # 批量修改
        results = manager.update_case_titles(cases)
        conn.commit()
        return jsonify(results), 200
    except Exception as e:
        logger.error(f"Failed to batch update case titles: {e}")
        return jsonify({'error': 'Failed to batch update case titles'}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/app/update', methods=['GET'])
def get_update_info():
    env_config.reload()
    return jsonify({
        "version": env_config.global_setting.version,
        "desc": "修复若干bug，提升体验",
        # "url": f"{env_config.global_setting.protocol}://{env_config.global_setting.domain}/app/update/{env_config.global_setting.app_name}"
        "url": f"https://patvs.lenovo.com/app/update/{env_config.global_setting.app_name}"
    })


@app.route('/update/plan/<int:plan_id>/models', methods=['POST'])
@token_required
def add_model_to_plan_api(current_user, current_userid, plan_id):
    """为测试计划添加机型"""
    data = request.get_json()
    model_name = data.get('model_name')
    if not model_name:
        return jsonify({
            "success": False,
            "message": "机型名称或者userId不能为空"
        }), 400
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        manager = TestCaseManager(conn, cursor)
        result = manager.add_model_to_plan(plan_id, model_name, current_userid)
        conn.commit()
        return jsonify(result), 200

    except Exception as e:
        conn.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@app.route('/update/plan/<int:plan_id>/models/<int:model_id>', methods=['DELETE'])
@token_required
def remove_model_from_plan_api(current_user, current_userid, plan_id, model_id):
    """从测试计划中删除机型"""
    conn = get_db_connection()
    cursor = conn.cursor()
    manager = TestCaseManager(conn, cursor)
    try:
        result = manager.remove_model_from_plan(plan_id, model_id, current_userid)
        conn.commit()
        return jsonify(result), 200

    except Exception as e:
        conn.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@app.route('/update/plan/<int:plan_id>/models/<int:model_id>', methods=['PUT'])
@token_required
def update_model_in_plan_api(current_user, current_userid, plan_id, model_id):
    """修改测试计划中的机型"""
    data = request.get_json()
    new_model_name = data.get('model_name')
    conn = get_db_connection()
    cursor = conn.cursor()
    manager = TestCaseManager(conn, cursor)
    try:
        if not new_model_name:
            return jsonify({
                "success": False,
                "message": "新机型名称不能为空"
            }), 400
        result = manager.update_model_in_plan(plan_id, model_id, new_model_name, current_userid)
        conn.commit()
        return jsonify(result), 200

    except Exception as e:
        conn.rollback()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


if __name__ == '__main__':
    app.run(debug=env_config.global_setting.is_debug, host=env_config.global_setting.domain,
            port=env_config.global_setting.port)
