# -*- coding: utf-8 -*-
# 负责GUI界面展示以及交互逻辑
import os
import json
import time
import itertools
import pywintypes
import wx
import wx.grid
import subprocess
import threading
import openpyxl
import sys
from functools import partial
from monitor_manager.patvs_fuction import Patvs_Fuction
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from monitor_manager.up_files import run_main
from datetime import datetime
from common.logs import logger
import win32con
import win32api
from requests_manager.http_requests_manager import http_manager
import wx.lib.newevent
import webbrowser


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


class ButtonRenderer(wx.grid.GridCellRenderer):
    """自定义按钮渲染器，仅显示蓝色字体"""

    def __init__(self, label="按钮"):
        super().__init__()
        self.label = label  # 按钮上的文字

    def Draw(self, grid, attr, dc, rect, row, col, isSelected):
        """绘制蓝色字体"""
        # 清除背景
        dc.SetBrush(wx.Brush(wx.Colour(255, 255, 255)))  # 白色背景
        dc.SetPen(wx.TRANSPARENT_PEN)  # 无边框
        dc.DrawRectangle(rect.x, rect.y, rect.width, rect.height)

        # 设置字体颜色为蓝色
        dc.SetTextForeground(wx.Colour(0, 0, 255))  # 蓝色字体
        font = dc.GetFont()
        font.SetWeight(wx.FONTWEIGHT_BOLD)  # 粗体文字
        dc.SetFont(font)

        # 绘制文字
        text_width, text_height = dc.GetTextExtent(self.label)
        text_x = rect.x + (rect.width - text_width) // 2
        text_y = rect.y + (rect.height - text_height) // 2
        dc.DrawText(self.label, text_x, text_y)

    def GetBestSize(self, grid, attr, dc, row, col):
        """返回按钮的最佳尺寸"""
        dc.SetFont(grid.GetFont())
        text_width, text_height = dc.GetTextExtent(self.label)
        # 返回文本的宽度和高度，并添加适当的边距
        return wx.Size(text_width + 10, text_height + 10)

    def Clone(self):
        """克隆渲染器实例"""
        return ButtonRenderer(self.label)


class TestCasesPanel(wx.Panel):
    def __init__(self, parent, username, token):
        super().__init__(parent)
        # 在主线程中创建一个事件,用来通知阻塞情况下终止线程
        self.stop_event = True
        self.patvs_monitor = Patvs_Fuction(self, self.stop_event)
        self.username = username  # 保存用户名
        self.userid = http_manager.get_params(f'/get_userid/{self.username}').get('user_id')  # 保存用户名
        self.token = token
        self.splitter = wx.SplitterWindow(self)
        # 初始化 log_splitter
        self.log_splitter = wx.SplitterWindow(self.splitter)
        self.tree = wx.TreeCtrl(self.splitter, style=wx.TR_DEFAULT_STYLE)
        self.content = wx.TextCtrl(self.log_splitter, style=wx.TE_MULTILINE | wx.TE_READONLY | wx.TE_RICH2)
        self.log_content = wx.TextCtrl(self.log_splitter, style=wx.TE_MULTILINE | wx.TE_READONLY | wx.TE_RICH2)

        # 初始化分隔条，确保位置设置稳定
        self.log_splitter.SetMinimumPaneSize(20)  # 设置最小窗格尺寸
        self.log_splitter.SplitHorizontally(self.content, self.log_content, sashPosition=-200)
        self.log_splitter.SetSashGravity(0.8)

        self.splitter.SetMinimumPaneSize(20)
        self.splitter.SplitVertically(self.tree, self.log_splitter, sashPosition=200)
        self.splitter.SetSashGravity(0.2)

        # 上传按钮
        upload_icon = wx.Bitmap(resource_path("icon\\上传文件夹.png"))
        self.upload_button = wx.BitmapButton(self, bitmap=upload_icon)
        self.upload_button.SetToolTip(wx.ToolTip("上传文件"))
        self.upload_button.Bind(wx.EVT_BUTTON, self.up_file)

        # 检查按钮
        device_icon = wx.Bitmap(resource_path("icon\\设备管理.png"))
        # self.check_button = wx.Button(self, label='检查')
        self.device_button = wx.BitmapButton(self, bitmap=device_icon)
        self.device_button.SetToolTip(wx.ToolTip("打开设备管理器"))
        self.device_button.Bind(wx.EVT_BUTTON, self.open_device_manager)

        # 配置按钮
        config_icon = wx.Bitmap(resource_path("icon\\configure.png"))
        self.config_button = wx.BitmapButton(self, bitmap=config_icon)
        self.config_button.SetToolTip(wx.ToolTip("生成配置文件"))
        self.config_button.Bind(wx.EVT_BUTTON, self.write_config)

        # 查看用例状态按钮
        status_icon = wx.Bitmap(resource_path("icon\\查看状态1.png"))
        self.status_button = wx.BitmapButton(self, bitmap=status_icon)
        self.status_button.SetToolTip(wx.ToolTip("查看用例状态"))
        self.status_button.Bind(wx.EVT_BUTTON, self.check_status)

        # 附件按钮
        annex_icon = wx.Bitmap(resource_path("icon\\game-icons--combination-lock.png"))
        self.annex_button = wx.BitmapButton(self, bitmap=annex_icon)
        self.annex_button.SetToolTip(wx.ToolTip("排列组合生成器"))
        self.annex_button.Bind(wx.EVT_BUTTON, self.permutation_and_combination)

        # 用例筛选下拉框
        # 固定宽度
        fixed_width = 200
        self.project_name_combo = wx.ComboBox(self)
        self.project_name_combo.SetMinSize(wx.Size(150, -1))
        self.project_name_combo.Bind(wx.EVT_COMBOBOX, self.on_project_select)
        self.project_name_combo.Bind(wx.EVT_MOTION, self.on_project_hover) # 鼠标悬停时动态加载数据

        self.plan_name_combo = wx.ComboBox(self)
        self.plan_name_combo.SetMinSize(wx.Size(fixed_width, -1))
        self.plan_name_combo.Bind(wx.EVT_COMBOBOX, self.on_plan_select)
        self.plan_name_combo.Bind(wx.EVT_MOTION, self.on_plan_hover)

        self.model_name_combo = wx.ComboBox(self)
        self.model_name_combo.SetMinSize(wx.Size(fixed_width, -1))
        self.model_name_combo.Bind(wx.EVT_COMBOBOX, self.on_model_select)
        self.model_name_combo.Bind(wx.EVT_MOTION, self.on_model_hover)

        self.sheet_name_combo = wx.ComboBox(self)
        self.sheet_name_combo.SetMinSize(wx.Size(fixed_width, -1))
        self.sheet_name_combo.Bind(wx.EVT_COMBOBOX, self.on_sheet_select)
        self.sheet_name_combo.Bind(wx.EVT_MOTION, self.on_sheet_hover)

        # 创建一个字体对象，字体大小为16，字体家族为瑞士，风格为正常，但是字体粗细为加粗
        font = wx.Font(12, wx.SWISS, wx.NORMAL, wx.NORMAL)

        # 创建布局，将按钮放置在左上角
        buttonSizer = wx.BoxSizer(wx.HORIZONTAL)
        labelSizer = wx.BoxSizer(wx.HORIZONTAL)

        # 将按钮添加到布局中
        buttonSizer.Add(self.upload_button, 0, wx.ALL, 5)
        buttonSizer.Add(self.device_button, 0, wx.ALL, 5)
        buttonSizer.Add(self.config_button, 0, wx.ALL, 5)
        buttonSizer.Add(self.status_button, 0, wx.ALL, 5)
        buttonSizer.Add(self.annex_button, 0, wx.ALL, 5)

        self.testerLabel = wx.StaticText(self, label=f"测试人员: {username}")
        labelSizer.Add(self.testerLabel, 0, wx.ALL, 8)

        self.test_phaseLabel = wx.StaticText(self, label=f"测试阶段: N/A")
        labelSizer.Add(self.test_phaseLabel, 0, wx.ALL, 8)

        # 添加统计的标签（初始为空）
        self.case_time_total = wx.StaticText(self, label="总耗时: N/A")
        #   self.case_time_total.SetFont(font)
        labelSizer.Add(self.case_time_total, 0, wx.ALL, 8)

        self.case_total = wx.StaticText(self, label="用例总数: N/A")
        #    self.case_total.SetFont(font)
        labelSizer.Add(self.case_total, 0, wx.ALIGN_CENTER | wx.ALL, 8)

        self.executed_cases = wx.StaticText(self, label="已执行用例: N/A")
        #  self.executed_cases.SetFont(font)
        labelSizer.Add(self.executed_cases, 0, wx.ALIGN_CENTER | wx.ALL, 8)

        self.pass_count = wx.StaticText(self, label="Pass: N/A")
        #    self.pass_count.SetFont(font)
        labelSizer.Add(self.pass_count, 0, wx.ALIGN_CENTER | wx.ALL, 8)

        self.fail_count = wx.StaticText(self, label="Fail: N/A")
        #   self.fail_count.SetFont(font)
        labelSizer.Add(self.fail_count, 0, wx.ALIGN_CENTER | wx.ALL, 8)

        self.block_count = wx.StaticText(self, label="Block: N/A")
        #  self.block_count.SetFont(font)
        labelSizer.Add(self.block_count, 0, wx.ALIGN_CENTER | wx.ALL, 8)

        # 创建主布局
        mainSizer = wx.BoxSizer(wx.VERTICAL)

        # 在主布局中加入按钮布局，并使其水平居中
        mainSizer.Add(buttonSizer, 0, wx.EXPAND)
        mainSizer.Add(labelSizer, 0, wx.EXPAND)

        # 单独创建一个水平盒子来放置 case_search_combo 下拉框
        caseSearchSizer = wx.BoxSizer(wx.HORIZONTAL)
        self.projectLabel = wx.StaticText(self, label="项目")
        caseSearchSizer.Add(self.projectLabel, 0, wx.ALL, 5)
        caseSearchSizer.Add(self.project_name_combo, 0, wx.ALL, 5)

        self.planLabel = wx.StaticText(self, label="测试计划")
        caseSearchSizer.Add(self.planLabel, 0, wx.ALL, 5)
        caseSearchSizer.Add(self.plan_name_combo, 0, wx.ALL, 5)

        self.modelLabel = wx.StaticText(self, label="测试机型")
        caseSearchSizer.Add(self.modelLabel, 0, wx.ALL, 5)
        caseSearchSizer.Add(self.model_name_combo, 0, wx.ALL, 5)

        self.sheetLabel = wx.StaticText(self, label="测试用例")
        caseSearchSizer.Add(self.sheetLabel, 0, wx.ALL, 5)
        caseSearchSizer.Add(self.sheet_name_combo, 0, wx.ALL, 5)

        # # 添加透明占位符,保持对齐
        # dummyLabel = wx.StaticText(self, label="")
        # dummyLabel.SetMinSize((60, -1))
        # caseSearchSizer.Add(dummyLabel, 0, wx.EXPAND)

        # 在主布局中加入 case_search_combo 布局，确保它在按钮下方
        mainSizer.Add(caseSearchSizer, 0, wx.EXPAND)

        # 创建新的布局，将下拉框和按钮放在右下脚
        actionSizer = wx.BoxSizer(wx.HORIZONTAL)
        buttonSizer2 = wx.BoxSizer(wx.HORIZONTAL)
        # 图片
        lenovo_bitmap = wx.Bitmap(resource_path("icon\\3332.png"), wx.BITMAP_TYPE_ANY)
        imageCtrl = wx.StaticBitmap(self, bitmap=lenovo_bitmap)

        # start、pass、fail、block 按钮
        self.start_button = wx.Button(self, label='start')
        self.start_button.Bind(wx.EVT_BUTTON, self.start_test)
        # 创建按钮并添加到布局
        self.result_buttons = {}
        for button in ['Pass', 'Fail', 'Block']:
            self.result_buttons[button] = wx.Button(self, label=button)
            self.result_buttons[button].Bind(wx.EVT_BUTTON, self.test_result)
            buttonSizer2.Add(self.result_buttons[button], 0, wx.ALL, 5)
            self.result_buttons[button].Hide()

        self.actions_and_num = wx.StaticText(self, label="监控动作和次数: N/A")

        # 在主布局中添加按钮布局和新的布局
        mainSizer.Add(self.splitter, 1, wx.EXPAND)

        actionSizer.Add(self.actions_and_num, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)
        actionSizer.AddSpacer(250)  # 添加伸缩控件

        # 创建底部布局，将下拉框和按钮放在右侧
        bottomSizer = wx.BoxSizer(wx.HORIZONTAL)
        # 添加图片到布局（最左边）
        bottomSizer.Add(imageCtrl, 0, wx.ALL | wx.ALIGN_LEFT, 5)
        bottomSizer.Add(actionSizer, 1, wx.EXPAND | wx.ALL, 5)
        bottomSizer.Add(self.start_button, 0, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 5)
        bottomSizer.Add(buttonSizer2, 0, wx.ALL, 5)
        mainSizer.Add(bottomSizer, 0, wx.ALL, 5)

        # 设置主布局
        self.SetSizer(mainSizer)
        # 初始时不显示用例树，等待用户选择计划和用例表后显示
        self._setup_icons()  # 初始化时设置图标列表
        self.tree.Hide()
        # 初始时没有用例树数据
        self.testCases = None
        self.plan_id = None
        self.model_id = None
        self.sheet_id = None
        self.CaseID = None
        self.tree.Bind(wx.EVT_TREE_SEL_CHANGED, self.case_details)
        # 如果有记录，显示上一次打开的页面
        self.restore_state()

    def _setup_icons(self):
        icons = {
            "Pass": resource_path("icon\\Pass.png"),
            "Fail": resource_path("icon\\Fail.png"),
            "Block": resource_path("icon\\Block.png"),
            "None": None,
            "Root": resource_path("icon\\rootIcon.png")
        }
        self.image_list = wx.ImageList(16, 16)
        self.icon_indices = {}
        transparent_bmp = wx.Bitmap(16, 16, 32)
        image = transparent_bmp.ConvertToImage()
        image.InitAlpha()
        for x in range(16):
            for y in range(16):
                image.SetAlpha(x, y, 0)
        transparent_bmp = wx.Bitmap(image)
        for status, icon in icons.items():
            if icon:
                self.icon_indices[status] = self.image_list.Add(wx.Bitmap(icon))
            else:
                self.icon_indices[status] = self.image_list.Add(transparent_bmp)

    def on_project_select(self, event):
        logger.warning("开始调用 on_project_select")
        # 获取选择的项目名称
        project_name = self.project_name_combo.GetValue()

        # 根据项目名称获取计划名称
        plan_names_with_ids = http_manager.get_plan_names(self.userid, project_name)

        # 清空并重新填充 plan_name_combo
        self.plan_name_combo.Clear()
        for plan_id, plan_name in plan_names_with_ids:
            self.plan_name_combo.Append(plan_name, plan_id)

        # 清空 sheet_name_combo，因为项目改变可能导致计划和用例表的变化
        self.model_name_combo.Clear()
        self.sheet_name_combo.Clear()
        self.tree.Hide()  # 隐藏用例树
        self.clear_statistics()  # 清空统计信息
        logger.warning("结束调用 on_project_select")

    def on_plan_select(self, event):
        # 选择测试计划后，更新用例表下拉框
        logger.warning('开始调用 plan select')
        selected_index = self.plan_name_combo.GetSelection()
        if selected_index != wx.NOT_FOUND:
            self.plan_id = self.plan_name_combo.GetClientData(selected_index)
            self.plan_name = self.plan_name_combo.GetString(selected_index)
            logger.warning(self.plan_id)
            logger.warning(self.plan_name)
        model_names_with_ids = http_manager.get_params(f'/get_model_names/{self.plan_id}').get('model_names')

        # 清空并重新填充 model_name_combo
        self.model_name_combo.Clear()
        self.sheet_name_combo.Clear()
        for model_id, model_name in model_names_with_ids:
            self.model_name_combo.Append(model_name, model_id)
        self.tree.Hide()  # 选择计划后先隐藏用例树，等待选择用例表后再显示
        # 清空统计信息
        self.clear_statistics()
        logger.warning('调用 plan select 结束')

    def on_model_select(self, event):
        # 选择测试计划后，更新测试机型
        logger.warning('开始调用 model select')
        selected_index = self.model_name_combo.GetSelection()
        if selected_index != wx.NOT_FOUND:
            self.model_id = self.model_name_combo.GetClientData(selected_index)
            self.model_name = self.model_name_combo.GetString(selected_index)
            logger.warning(self.model_id)
            logger.warning(self.model_name)
        sheet_names_with_ids = http_manager.get_sheet_names(self.plan_id)
        # 清空并重新填充 sheet_name_combo
        self.sheet_name_combo.Clear()
        for sheet_id, sheet_name in sheet_names_with_ids:
            self.sheet_name_combo.Append(sheet_name, sheet_id)
        self.tree.Hide()  # 选择计划后先隐藏用例树，等待选择用例表后再显示
        # 清空统计信息
        self.clear_statistics()
        logger.warning('调用 model select 结束')

    def on_sheet_select(self, event):
        # 选择用例表后，更新用例树
        selected_index = self.sheet_name_combo.GetSelection()
        if selected_index != wx.NOT_FOUND:
            self.sheet_id = self.sheet_name_combo.GetClientData(selected_index)
            self.sheet_name = self.sheet_name_combo.GetString(selected_index)
            self.case_tree(self.sheet_id, self.model_id)  # 使用 sheet_id 获取用例树
            self.tree.Show()  # 选择用例表后显示用例树
            # 更新统计信息
            self.update_statistics()
            self.CaseID = None

    def on_project_hover(self, event):
        """
        当鼠标悬停在项目下拉框时，动态加载项目数据，同时保留当前选中的项目
        """
        logger.info("开始动态加载项目数据")
        try:
            # 获取当前选中的项目名称
            current_selection = self.project_name_combo.GetValue()

            # 调用接口获取最新的项目名称列表
            project_names = http_manager.get_params(f'/get_project_names/{self.userid}').get('project_names', [])

            # 清空并重新填充 project_name_combo
            self.project_name_combo.Clear()
            for project_name in project_names:
                self.project_name_combo.Append(project_name)

            # 如果之前有选中的项目，尝试重新设置选中状态
            if current_selection in project_names:
                self.project_name_combo.SetValue(current_selection)
            else:
                # 如果当前选中的项目不在最新列表中，清空选中状态
                self.project_name_combo.SetValue("")

            logger.info("项目数据加载完成")
        except Exception as e:
            logger.error(f"加载项目数据失败: {e}")
        event.Skip()

    def on_plan_hover(self, event):
        # 获取当前选择的内容
        selection = self.plan_name_combo.GetStringSelection()
        if selection:
            self.plan_name_combo.SetToolTip(selection)
        event.Skip()

    def on_model_hover(self, event):
        # 获取当前选择的内容
        selection = self.model_name_combo.GetStringSelection()
        if selection:
            self.model_name_combo.SetToolTip(selection)
        event.Skip()

    def on_sheet_hover(self, event):
        # 获取当前选择的内容
        selection = self.sheet_name_combo.GetStringSelection()
        if selection:
            self.sheet_name_combo.SetToolTip(selection)
        event.Skip()

    def clear_statistics(self):
        # 清空统计信息
        self.test_phaseLabel.SetLabel("测试阶段: N/A")
        self.case_time_total.SetLabel("总耗时: N/A")
        self.case_total.SetLabel("用例总数: N/A")
        self.executed_cases.SetLabel("已执行用例: N/A")
        self.pass_count.SetLabel("Pass: N/A")
        self.fail_count.SetLabel("Fail: N/A")
        self.block_count.SetLabel("Block: N/A")

    def update_statistics(self):
        # 更新统计信息
        if self.sheet_id:
            pamars = {'planId': self.plan_id, 'modelId': self.model_id, 'sheetId': self.sheet_id}
            data = http_manager.get_params(f'/calculate_progress_and_pass_rate', params=pamars)
            result = data.get('result')
            self.test_phaseLabel.SetLabel(f"测试阶段: {result['project_phase']}")
            self.case_time_total.SetLabel(f"总耗时: {result['case_time_count']}(Min)")
            self.case_total.SetLabel(f"用例总数: {result['case_count']}")
            self.executed_cases.SetLabel(f"已执行用例: {result['executed_cases_count']}")
            self.pass_count.SetLabel(f"Pass: {result['pass_count']}")
            self.fail_count.SetLabel(f"Fail: {result['fail_count']}")
            self.block_count.SetLabel(f"Block: {result['block_count']}")

    def save_state(self):
        state = {
            'username': self.username,
            'test_project': self.project_name_combo.GetValue() if self.project_name_combo.GetValue() else None,
            'test_plan': self.plan_name_combo.GetValue() if self.plan_name_combo.GetValue() else None,
            'plan_id': self.plan_id,
            'test_model': self.model_name_combo.GetValue() if self.model_name_combo.GetValue() else None,
            'model_id': self.model_id,
            'test_sheet': self.sheet_name_combo.GetValue() if self.sheet_name_combo.GetValue() else None,
            'sheet_id': self.sheet_id,
            'test_case_id': self.tree.GetItemData(
                self.tree.GetSelection()) if self.tree.GetSelection().IsOk() else None,
            'start_clicked': self.start_clicked if hasattr(self, 'start_clicked') else False
        }
        logger.info('Saving state.')
        with open(r'C:\PATVS\window_state.json', 'w') as state_file:
            json.dump(state, state_file)

    def restore_state(self):
        try:
            with open(r'C:\PATVS\window_state.json', 'r') as state_file:
                state = json.load(state_file)
                # 防止 username 篡改数据
                if state['username'] == self.username:
                    if 'test_project' in state and state['test_project']:
                        self.project_name_combo.SetValue(state['test_project'])
                        self.on_project_select(None)
                    if 'test_plan' in state and state['test_plan']:
                        self.plan_name_combo.SetValue(state['test_plan'])
                        self.plan_id = state.get('plan_id', None)
                        self.on_plan_select(None)
                    if 'test_model' in state and state['test_model']:
                        self.model_name_combo.SetValue(state['test_model'])
                        self.model_id = state.get('model_id', None)
                        self.model_name = state.get('test_model', None)
                        self.on_model_select(None)
                    if 'test_sheet' in state and state['test_sheet']:
                        # self.sheet_name_combo.SetValue(state['test_sheet'])
                        # self.sheet_id = state.get('sheet_id', None)
                        # self.sheet_name = state.get('test_sheet', None)
                        # self.on_sheet_select(None)
                        sheet_name = state['test_sheet']
                        sheet_id = state.get('sheet_id', None)
                        sheet_names_with_ids = http_manager.get_sheet_names(self.plan_id)

                        # 清空并重新填充 sheet_name_combo
                        self.sheet_name_combo.Clear()
                        for sid, sname in sheet_names_with_ids:
                            self.sheet_name_combo.Append(sname, sid)
                            # 如果找到匹配的 sheet_name，设置选项和客户端数据
                            if sname == sheet_name and sid == sheet_id:
                                selected_index = self.sheet_name_combo.FindString(sheet_name)
                                if selected_index != wx.NOT_FOUND:
                                    self.sheet_name_combo.SetSelection(selected_index)
                                    self.sheet_name_combo.SetClientData(selected_index, sheet_id)
                                    logger.info(f"Restored sheet_name: {sheet_name}, sheet_id: {sheet_id}")
                                    self.on_sheet_select(None)  # 在找到匹配的sheet_name和sheet_id时调用
                    if 'test_case_id' in state and state['test_case_id']:
                        self.set_tree_selection_by_id(state['test_case_id'])
                    if 'start_clicked' in state and state['start_clicked']:
                        self.start_clicked = True
                        self.start_test(None)
                    else:
                        self.start_clicked = False
        except FileNotFoundError:
            logger.info('State file not found, starting with default state.')
            self.Center()

    def set_tree_selection_by_id(self, case_id):
        root = self.tree.GetRootItem()
        if root:
            self.expand_and_select_by_id(root, case_id, [])

    def expand_and_select_by_id(self, item, case_id, path):
        if self.tree.GetItemData(item) == case_id:
            for p_item in path:
                self.tree.Expand(p_item)
            self.tree.SelectItem(item)
            wx.CallAfter(self.tree.EnsureVisible, item)  # Make sure the item is visible
            wx.CallAfter(self.tree.SetFocus)  # Set focus to the tree
            return True

        path.append(item)
        child, cookie = self.tree.GetFirstChild(item)
        while child.IsOk():
            if self.expand_and_select_by_id(child, case_id, path):
                return True
            child, cookie = self.tree.GetNextChild(item, cookie)
        path.pop()
        return False

    def add_log_message(self, message):
        """向日志窗口添加消息"""
        if self.log_content:
            self.log_content.AppendText(message + '\n')  # 在文本控件的末尾添加文本
            logger.info(message + '\n')

    def upload_image(self, case_result, comment=None):
        """改进后的上传图片交互逻辑，支持图片预览和删除功能"""
        dialog = wx.Dialog(self, title="上传图片", size=(800, 600))

        # 主布局
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        # 图片预览区域
        preview_label = wx.StaticText(dialog, label="已选择的图片（最多 5 张）：")
        main_sizer.Add(preview_label, 0, wx.ALL, 5)

        # 使用 WrapSizer 动态排列图片预览
        self.image_preview_sizer = wx.WrapSizer(wx.HORIZONTAL)
        preview_panel = wx.Panel(dialog)
        preview_panel.SetSizer(self.image_preview_sizer)
        main_sizer.Add(preview_panel, 1, wx.ALL | wx.EXPAND, 5)

        # 按钮区域
        button_sizer = wx.BoxSizer(wx.HORIZONTAL)
        add_button = wx.Button(dialog, label="添加图片")
        confirm_button = wx.Button(dialog, label="确认上传")
        cancel_button = wx.Button(dialog, label="取消")
        button_sizer.Add(add_button, 0, wx.ALL, 5)
        button_sizer.AddStretchSpacer()
        button_sizer.Add(confirm_button, 0, wx.ALL, 5)
        button_sizer.Add(cancel_button, 0, wx.ALL, 5)
        main_sizer.Add(button_sizer, 0, wx.EXPAND)

        dialog.SetSizer(main_sizer)

        # 初始化已选图片列表
        selected_files = []

        def refresh_preview():
            """刷新图片预览区域"""
            # 清空现有预览
            for child in self.image_preview_sizer.GetChildren():
                child.GetWindow().Destroy()
            self.image_preview_sizer.Clear()

            # 添加每张图片的预览
            for file_path in selected_files:
                # 创建图片预览框
                image_panel = wx.Panel(preview_panel, size=(150, 200))  # 每个图片预览框的固定大小
                image_sizer = wx.BoxSizer(wx.VERTICAL)

                # 加载图片并显示缩略图
                image = wx.Image(file_path, wx.BITMAP_TYPE_ANY)
                image = image.Scale(120, 120, wx.IMAGE_QUALITY_HIGH)  # 缩略图大小
                bitmap = wx.StaticBitmap(image_panel, bitmap=wx.Bitmap(image))
                image_sizer.Add(bitmap, 0, wx.ALL | wx.CENTER, 5)

                # 显示文件名
                filename = os.path.basename(file_path)
                filename_label = wx.StaticText(image_panel, label=filename, style=wx.ALIGN_CENTER)
                image_sizer.Add(filename_label, 0, wx.ALL | wx.CENTER, 5)

                # 删除按钮
                delete_button = wx.Button(image_panel, label="删除", size=(60, 30))
                delete_button.Bind(wx.EVT_BUTTON, lambda evt, path=file_path: on_remove_image(evt, path))
                image_sizer.Add(delete_button, 0, wx.ALL | wx.CENTER, 5)

                image_panel.SetSizer(image_sizer)
                self.image_preview_sizer.Add(image_panel, 0, wx.ALL, 10)

            # 更新布局
            preview_panel.Layout()
            dialog.Layout()

        def on_add_image(event):
            """添加图片"""
            with wx.FileDialog(dialog, "选择图片文件",
                               wildcard="JPEG files (*.jpg;*.jpeg)|*.jpg;*.jpeg|PNG files (*.png)|*.png",
                               style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST | wx.FD_MULTIPLE) as fileDialog:

                if fileDialog.ShowModal() == wx.ID_CANCEL:
                    return  # 用户取消操作

                # 获取选择的文件路径
                paths = fileDialog.GetPaths()
                for path in paths:
                    if path not in selected_files:
                        if len(selected_files) >= 5:
                            wx.MessageBox('最多只能上传 5 张图片。', '提示', wx.OK | wx.ICON_WARNING)
                            break
                        selected_files.append(path)

                refresh_preview()

        def on_remove_image(event, file_path):
            """删除选中的图片"""
            if file_path in selected_files:
                selected_files.remove(file_path)
                refresh_preview()  # 删除后刷新预览区域

        def on_confirm_upload(event):
            """确认上传"""
            if not selected_files:
                wx.MessageBox('请选择至少一张图片进行上传。', '提示', wx.OK | wx.ICON_WARNING)
                return

            try:
                # 调用上传接口
                self.upload_files_to_server(selected_files, case_result, comment)
             #   wx.MessageBox('图片上传成功！', '信息', wx.OK | wx.ICON_INFORMATION)
                dialog.EndModal(wx.ID_OK)  # 上传成功后关闭对话框
            except Exception as e:
                wx.MessageBox(f"图片上传失败：{e}", '错误', wx.OK | wx.ICON_ERROR)

        def on_cancel(event):
            """取消操作"""
            dialog.EndModal(wx.ID_CANCEL)

        # 绑定按钮事件
        add_button.Bind(wx.EVT_BUTTON, on_add_image)
        confirm_button.Bind(wx.EVT_BUTTON, on_confirm_upload)
        cancel_button.Bind(wx.EVT_BUTTON, on_cancel)

        # 显示对话框并等待用户操作
        result = dialog.ShowModal()
        dialog.Destroy()

        # 如果用户点击确认上传返回 True，否则返回 False
        return result == wx.ID_OK

    def upload_files_to_server(self, file_paths, case_result, comment=None):
        files = []
        for file_path in file_paths:
            # 打开文件并保持打开状态，直到上传完成
            file = open(file_path, 'rb')
            files.append(('image_files', (os.path.basename(file_path), file, 'multipart/form-data')))

        data = {'case_id': self.CaseID, 'model_id': self.model_id, 'case_result': case_result, 'comment': comment}
        logger.warning(files)
        try:
            response = http_manager.post_file('/upload-images', files=files, data=data, token=self.token)
            if response.status_code == 200:
                wx.MessageBox('图片上传成功！', '信息', wx.OK | wx.ICON_INFORMATION)
            else:
                wx.MessageBox('图片上传失败！', '错误', wx.OK | wx.ICON_ERROR)
        finally:
            # 确保所有文件都被关闭
            for _, (filename, file, _) in files:
                file.close()

    def test_result(self, event):
        clicked_button = event.GetEventObject()
        case_result = clicked_button.GetLabel()
        action_and_num = http_manager.get_params(f'/get_case_actions_and_num/{self.CaseID}').get('actions_and_num')
        if clicked_button is self.result_buttons['Pass']:
            # 处理 pass
            logger.info(f"Pass Button clicked {self.CaseID}")
            if len(action_and_num) == 1 and '时间' in action_and_num[0]:
                # 主观测试强制用户上传图片
                image_uploaded = self.upload_image(case_result)
                if not image_uploaded:
                    wx.MessageBox('上传图片是必填操作，请上传图片后再继续。', '提示', wx.OK | wx.ICON_WARNING)
                    return  # 如果用户没有上传图片，则返回，阻止后续操作
            else:
                http_manager.update_end_time_case_id(self.CaseID, self.model_id, case_result, token=self.token)
            wx.CallAfter(self.case_enable)
            wx.CallAfter(self.refresh_node_case_status, case_status=case_result)
            wx.CallAfter(self.update_statistics)
        if clicked_button is self.result_buttons['Fail']:
            # 处理 Fail
            # 弹出文本输入对话框
            dlg = wx.TextEntryDialog(self, '请输入Fail原因 :', 'Comment')
            if dlg.ShowModal() == wx.ID_OK:
                input_content = dlg.GetValue().strip()  # 获取输入的内容
                if input_content:
                    if len(action_and_num) == 1 and '时间' in action_and_num[0]:
                        # 主观测试强制用户上传图片
                        image_uploaded = self.upload_image(case_result, f'Fail: {input_content}')
                        if not image_uploaded:
                            wx.MessageBox('上传图片是必填操作，请上传图片后再继续。', '提示', wx.OK | wx.ICON_WARNING)
                            return  # 如果用户没有上传图片，则返回，阻止后续操作
                    else:
                        http_manager.update_end_time_case_id(self.CaseID, self.model_id, case_result,
                                                             f'Fail: {input_content}', self.token)
                    logger.info(f"Fail Button clicked, Content: {input_content}")

                    # 设置事件以通知监控线程停止
                    self.patvs_monitor.stop_event = False
                    # 当用例为 block 时，需要主动去停止 messageLoop 的循环
                    if self.patvs_monitor.msg_loop_thread_id:
                        logger.warning("进入终止消息循环")
                        try:
                            win32api.PostThreadMessage(self.patvs_monitor.msg_loop_thread_id, win32con.WM_QUIT, 0, 0)
                        except pywintypes.error as e:
                            logger.warning(f"{e}")
                        except:
                            pass
                        self.patvs_monitor.msg_loop_thread_id = None
                    time.sleep(1)  # block后线程终止需要一些时间，防止出现意外解禁按钮
                    wx.CallAfter(self.case_enable)
                    wx.CallAfter(self.refresh_node_case_status, case_status=case_result)
                    wx.CallAfter(self.update_statistics)
            else:
                wx.MessageDialog(self, '内容不能为空，请重新输入!', '错误', style=wx.OK | wx.ICON_ERROR).ShowModal()
            dlg.Destroy()
        if clicked_button is self.result_buttons['Block']:
            # 处理 fail or block
            # 弹出文本输入对话框
            dlg = wx.TextEntryDialog(self, '请输入Block原因 :', 'Comment')
            if dlg.ShowModal() == wx.ID_OK:
                input_content = dlg.GetValue().strip()  # 获取输入的内容
                if input_content:
                    logger.info(f"Block Button clicked, Content: {input_content}")
                    http_manager.update_end_time_case_id(self.CaseID, self.model_id, case_result,
                                                         f'Block: {input_content}',
                                                         self.token)
                    # 设置事件以通知监控线程停止
                    self.patvs_monitor.stop_event = False
                    # 当用例为 block 时，需要主动去停止 messageLoop 的循环
                    if self.patvs_monitor.msg_loop_thread_id:
                        logger.warning("进入终止消息循环")
                        try:
                            win32api.PostThreadMessage(self.patvs_monitor.msg_loop_thread_id, win32con.WM_QUIT, 0, 0)
                        except pywintypes.error as e:
                            logger.warning(f"{e}")
                        except:
                            pass
                        self.patvs_monitor.msg_loop_thread_id = None
                    time.sleep(1)  # block后线程终止需要一些时间，防止出现意外解禁按钮
                    wx.CallAfter(self.case_enable)
                    wx.CallAfter(self.refresh_node_case_status, case_status=case_result)
                    wx.CallAfter(self.update_statistics)
                else:
                    wx.MessageDialog(self, '内容不能为空，请重新输入!', '错误', style=wx.OK | wx.ICON_ERROR).ShowModal()
            dlg.Destroy()
        self.start_clicked = False

    def start_test(self, event):
        try:
            self.start_clicked = True
            # 检查是否有选中的用例
            if not hasattr(self, 'CaseID') or not self.CaseID:
                wx.MessageBox('请先选择用例', 'Warning')
                return
            action_and_num = http_manager.get_params(f'/get_case_actions_and_num/{self.CaseID}').get('actions_and_num')
            logger.warning(action_and_num)
            if not action_and_num:
                wx.MessageBox('未检测到任何匹配项，请按照规则修改用例标题后再测试', 'Warning')
                return
            params = {"case_id": self.CaseID, "model_id": self.model_id}
            result = http_manager.get_params(f'/get_case_result', params=params)
            if result.get('case_result'):
                wx.MessageBox('已有测试结果，请重置此条测试用例后再进行测试', 'Warning')
                return
            wx.CallAfter(self.case_disable)
            start_time = http_manager.get_params(f'/get_start_time/{self.model_id}/{self.CaseID}').get('start_time')
            if not start_time:
                logger.warning('没有找到执行记录，插入开始时间')
                now = datetime.now()
                start_time = now.strftime('%Y-%m-%d %H:%M:%S')
                http_manager.post_data('/update_start_time',
                                       {'case_id': self.CaseID, 'model_id': self.model_id, 'start_time': start_time},
                                       token=self.token)
            # 初始化终止信号
            self.patvs_monitor.stop_event = True

            # 使用多线程异步运行，防止GUI界面卡死
            thread = threading.Thread(target=self.patvs_monitor.run_main,
                                      args=(int(self.CaseID), list(action_and_num), str(start_time),))
            thread.setDaemon(True)
            thread.start()
            # 获取线程ID
            self.msg_loop_thread_id = thread.ident
        except Exception as e:
            wx.MessageBox(f'出现未知错误: {e},请联系系统管理员', 'Error')
            logger.error(f'出现未知错误: {e}')
            return

    def after_test(self):
        for button in self.result_buttons:
            self.result_buttons[button].Enable()
        self.Layout()

    def case_disable(self):
        """
        case 执行状态按钮显示
        :return:
        """
        self.status_button.Disable()
        self.plan_name_combo.Disable()
        self.sheet_name_combo.Disable()
        self.tree.Disable()
        self.start_button.Hide()
        for button in self.result_buttons:
            self.result_buttons[button].Show()
        self.result_buttons['Pass'].Disable()
        #  self.result_buttons['Fail'].Disable()
        self.Layout()

    def case_enable(self):
        """
        case 未执行状态按钮显示
        :return:
        """
        self.status_button.Enable()
        self.tree.Enable()
        self.plan_name_combo.Enable()
        self.sheet_name_combo.Enable()
        self.start_button.Show()
        for button in self.result_buttons:
            self.result_buttons[button].Hide()
        self.Layout()

    def refresh_node_case_status(self, case_id=None, case_status=None):
        root = self.tree.GetRootItem()
        if case_id is None:
            case_id = self.CaseID
        if not root.IsOk():
            return
        # 遍历所有的子节点
        (item, cookie) = self.tree.GetFirstChild(root)
        while item:
            # 如果当前节点的ID匹配，那么更新这个节点的图标
            if self.tree.GetItemData(item) == case_id:
                # 重置时 status 为 None
                if case_status is not None:
                    self.tree.SetItemImage(item, self.icon_indices[case_status])
                else:
                    self.tree.SetItemImage(item, self.icon_indices['None'])
                break
            # 获取下一个节点
            (item, cookie) = self.tree.GetNextChild(root, cookie)

    def up_file(self, event):
        # 打开文件对话框
        with wx.FileDialog(self, "选择测试用例文件", wildcard="文本文件 (*.xlsx)|*.xlsx",
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            # 获取选择的文件路径
            pathname = fileDialog.GetPath()
            filename = os.path.basename(pathname)  # 获取文件名
            # 校验文件是否已存在
            result = http_manager.get_params(f'/get_filename/{filename}').get('file_exists')
            logger.warning(f'file_exists: {result}')
            if result:
                wx.MessageBox(f"当前文件已存在，请勿重复上传", "提示", wx.OK | wx.ICON_WARNING)
                return

            # 显示等待提示框
            wait_dialog = wx.BusyInfo("正在上传和解析文件，请稍候...", self)

            def complete_upload():
                try:
                    run_main(pathname, self.userid, self.token)
                    wx.CallAfter(wx.MessageBox, "文件上传和解析成功", "提示", wx.OK | wx.ICON_INFORMATION)
                except Exception as e:
                    # 格式校验出错
                    wx.CallAfter(wx.MessageBox, f"上传的用例格式不符合规则，请按照模板导入用例。错误详情: {e}", "提示",
                                 wx.OK | wx.ICON_WARNING)
                    # 关闭等待提示框
                    wx.CallAfter(lambda: wait_dialog.__exit__(None, None, None))
                    return
                finally:
                    # 关闭等待提示框
                    wx.CallAfter(lambda: wait_dialog.__exit__(None, None, None))

                # # 更新显示
                # all_plans = http_manager.get_plan_names(self.username)
                # self.tree.DeleteAllItems()  # 清空现有的树状结构
                # self.plan_name_combo.Clear()
                # self.sheet_name_combo.Clear()  # 先清除之前的选项
                # self.plan_name_combo.AppendItems(all_plans)  # 添加新的选项
                # self.plan_name_combo.SetValue(all_plans[0])
                # self.on_plan_select(None)  # 自动加载 test_sheet

            wx.CallAfter(complete_upload)

    def open_device_manager(self, event):
        """
        打开Windows设备管理器
        :param event:
        :return:
        """

        def run_command():
            subprocess.run('mmc devmgmt.msc', shell=True)

        threading.Thread(target=run_command).start()

    def write_config(self, event):
        """
        写入配置信息/暂时不知道具体用途
        :param event:
        :return:
        """
        # 创建一个对话框用于输入配置信息
        dialog = wx.Dialog(None, title="配置信息", size=(350, 350))
        panel = wx.Panel(dialog)
        vbox = wx.BoxSizer(wx.VERTICAL)

        # 创建 BIOS,ECFW,MEFW, System SN,Retimer 的输入项
        labels = ["BIOS", "ECFW", "MEFW", "System SN", "Retimer"]
        self.inputs = {}
        for label in labels:
            hbox = wx.BoxSizer(wx.HORIZONTAL)
            st = wx.StaticText(panel, label=label)
            tc = wx.TextCtrl(panel)
            hbox.Add(st)
            hbox.Add(tc, flag=wx.LEFT, border=5)
            vbox.Add(hbox, flag=wx.EXPAND | wx.ALL, border=10)
            self.inputs[label] = tc

        # 为确认和取消按钮创建一行
        hbox = wx.BoxSizer(wx.HORIZONTAL)
        okButton = wx.Button(panel, label='保存')
        closeButton = wx.Button(panel, label='取消')

        # 绑定按钮事件
        okButton.Bind(wx.EVT_BUTTON, partial(self.save_config, dialog=dialog))
        closeButton.Bind(wx.EVT_BUTTON, partial(self.close_dialog, dialog=dialog))

        hbox.Add(okButton)
        hbox.Add(closeButton, flag=wx.LEFT, border=5)
        vbox.Add(hbox, flag=wx.ALIGN_CENTER | wx.TOP | wx.BOTTOM, border=10)

        panel.SetSizer(vbox)
        dialog.ShowModal()
        dialog.Destroy()

    def save_config(self, event, dialog):
        with open('config.txt', 'w') as f:
            for label, tc in self.inputs.items():
                f.write(f"{label}: {tc.GetValue()}\n")
        dialog.Close()
        wx.MessageBox("配置文件保存成功！", "信息", wx.OK | wx.ICON_INFORMATION)

    def close_dialog(self, event, dialog):
        dialog.Close()

    def check_status(self, event):
        if not self.sheet_id or not self.model_id:
            wx.MessageBox('请先选择用例', 'Warning')
            return

        # 获取用例数据
        self.testCases = http_manager.get_cases_by_sheet_id(self.sheet_id, self.model_id)

        # 创建一个新的对话框，并且允许对话框最大化
        dialog = wx.Dialog(self, title="查看用例状态", size=(800, 600), style=wx.MAXIMIZE_BOX | wx.DEFAULT_DIALOG_STYLE)

        # 创建grid并设置行和列
        self.grid = wx.grid.Grid(dialog)
        self.grid.CreateGrid(numRows=len(self.testCases), numCols=13)  # 增加两列用于按钮

        # 设置列标题
        cols_title = ['重置按钮', '测试结果', '测试耗时(S)', '用例标题', '前置条件',
                      '用例步骤', '预期结果', '开始时间', '完成时间', '评论', '失败次数', '阻塞次数', '查看图片']
        for i, title in enumerate(cols_title):
            self.grid.SetColLabelValue(i, title)

        # 填充数据
        self.row_to_execution_id = {}
        for i, case in enumerate(self.testCases):
            execution_id = case.get('ExecutionID')
            if execution_id is not None:
                self.row_to_execution_id[i] = execution_id  # 记录行号与 ExecutionID 的映射

            # 根据列标题填充数据
            self.grid.SetCellValue(i, 1, str(case.get('TestResult', "")))
            self.grid.SetCellValue(i, 2, str(case.get('TestTime', "")))
            self.grid.SetCellValue(i, 3, case.get('CaseTitle', ""))
            self.grid.SetCellValue(i, 4, str(case.get('PreConditions', "")))
            self.grid.SetCellValue(i, 5, case.get('CaseSteps', ""))
            self.grid.SetCellValue(i, 6, case.get('ExpectedResult', ""))
            self.grid.SetCellValue(i, 7, str(case.get('StartTime', "") or ""))
            self.grid.SetCellValue(i, 8, str(case.get('EndTime', "") or ""))
            self.grid.SetCellValue(i, 9, str(case.get('Comment', "") or ""))
            self.grid.SetCellValue(i, 10, str(case.get('FailCount', "") or ""))
            self.grid.SetCellValue(i, 11, str(case.get('BlockConut', "") or ""))

            # 设置背景颜色
            test_result = case.get('TestResult', "")
            if test_result == 'Pass':
                self.grid.SetCellBackgroundColour(i, 1, wx.Colour(144, 238, 144))  # 浅绿色
            elif test_result in ['Fail', 'Block']:
                self.grid.SetCellBackgroundColour(i, 1, wx.Colour(255, 99, 71))  # 浅红色

            # 为第一列设置“重置按钮”
            self.grid.SetCellValue(i, 0, "重置")
            self.grid.SetCellRenderer(i, 0, ButtonRenderer("重置"))

            # 为最后一列设置“查看图片按钮”
            self.grid.SetCellValue(i, 12, "查看图片")
            self.grid.SetCellRenderer(i, 12, ButtonRenderer("查看图片"))

        # 禁止直接编辑单元格
        self.grid.EnableEditing(False)

        # 绑定单元格点击事件
        self.grid.Bind(wx.grid.EVT_GRID_CELL_LEFT_CLICK, self.on_cell_click)
        # 自动调整每一列和每一行的大小以适应内容
        self.grid.AutoSizeColumns()
        self.grid.AutoSizeRows()

        # 创建下载按钮
        download_button = wx.Button(dialog, label="下载")
        download_button.Bind(wx.EVT_BUTTON, lambda evt: self.on_download(self.grid, evt))

        # 创建对话框的布局管理器，并将grid添加到其中
        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(self.grid, 1, wx.EXPAND | wx.ALL)
        sizer.Add(download_button, 0, wx.ALIGN_CENTER | wx.ALL, 10)
        dialog.SetSizer(sizer)

        # 显示对话框
        dialog.ShowModal()
        dialog.Destroy()

    def on_cell_click(self, event):
        """处理单元格点击事件"""
        row = event.GetRow()
        col = event.GetCol()

        # 判断点击的是“重置按钮”列
        if col == 0:
            execution_id = self.row_to_execution_id.get(row)
            if execution_id:
                self.on_reset_click(event, execution_id)
        # 判断点击的是“查看图片”列
        elif col == 12:
            execution_id = self.row_to_execution_id.get(row)
            if execution_id:
                self.on_view_image_click(execution_id)

        # 继续处理其他事件
        event.Skip()

    def reset_grid(self):
        self.grid.ClearGrid()
        self.testCases = http_manager.get_cases_by_sheet_id(self.sheet_id, self.model_id)

        for i, case in enumerate(self.testCases):
            self.grid.SetCellValue(i, 1, str(case.get('TestResult', "")))
            self.grid.SetCellValue(i, 2, str(case.get('TestTime', "")))
            self.grid.SetCellValue(i, 3, case.get('CaseTitle', ""))
            self.grid.SetCellValue(i, 4, str(case.get('PreConditions', "")))
            self.grid.SetCellValue(i, 5, case.get('CaseSteps', ""))
            self.grid.SetCellValue(i, 6, case.get('ExpectedResult', ""))
            self.grid.SetCellValue(i, 7, str(case.get('Actions', "")))
            self.grid.SetCellValue(i, 8, str(case.get('StartTime', "") or ""))
            self.grid.SetCellValue(i, 9, str(case.get('EndTime', "") or ""))
            self.grid.SetCellValue(i, 10, str(case.get('Comment', "") or ""))
            self.grid.SetCellValue(i, 11, str(case.get('FailCount', "") or ""))
            self.grid.SetCellValue(i, 12, str(case.get('BlockConut', "") or ""))

        self.grid.ForceRefresh()

    def on_reset_click(self, evt, execution_id):
        #  row, col = evt.GetRow(), evt.GetCol()
        """处理重置按钮点击事件"""
        case_result = http_manager.get_params(f'/get_case_result', params={"execution_id": execution_id}).get('case_result')
        if case_result:
            case_id = None
            for case in self.testCases:
                if case.get('ExecutionID') == execution_id:
                    case_id = case.get('CaseID')
                    break
            reset_msg = wx.MessageDialog(self, '您确定要重置这条用例测试结果吗?', '确认',
                                         wx.YES_NO | wx.ICON_QUESTION)
            reset_response = reset_msg.ShowModal()
            if reset_response == wx.ID_YES:
                logger.info(f"重置ExecutionID: {execution_id} 的用例测试状态")
                http_manager.post_data('/reset_case_result', {'execution_id': execution_id}, self.token)
                evt.Skip(False)
                self.reset_grid()  # 刷新网格布局
                wx.CallAfter(self.refresh_node_case_status, case_id=case_id)
                wx.CallAfter(self.update_statistics)
                self.tree.Refresh()
            reset_msg.Destroy()
        else:
            wx.MessageBox('此用例尚未执行，无法重置状态', '信息', wx.OK | wx.ICON_INFORMATION)
            return
        evt.Skip(True)

    def on_view_image_click(self, execution_id):
        """查看与 ExecutionID 相关的图片，生成 HTML 文件展示图片和信息"""

        # 调用接口获取图片
        response = http_manager.get_params(f'/get_images/{execution_id}')
        images = response.get('images', [])
        if not images:
            logger.warning(f"No images found for Execution ID {execution_id}: {response}")
            wx.MessageBox('未找到相关图片', 'Info')
            return

        # 生成 HTML 内容
        html_content = f"""
        <!DOCTYPE html>
        <html lang="zh-CN">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Execution ID {execution_id} 图片预览</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    margin: 20px;
                    line-height: 1.6;
                }}
                h1 {{
                    text-align: center;
                    color: #333;
                }}
                .image-card {{
                    margin-bottom: 30px;
                    border: 1px solid #ddd;
                    padding: 15px;
                    border-radius: 8px;
                    box-shadow: 0 2px 5px rgba(0, 0, 0, 0.1);
                }}
                .image-info {{
                    margin-bottom: 15px;
                }}
                .image-info p {{
                    margin: 5px 0;
                }}
                img {{
                    display: block;
                    max-width: 100%;
                    height: auto;
                    margin: 0 auto;
                    border: 1px solid #ccc;
                    border-radius: 5px;
                }}
            </style>
        </head>
        <body>
            <h1>Execution ID {execution_id} 图片预览</h1>
        """

        for index, image in enumerate(images):
            url = image.get('url')
            file_name = image.get('original_file_name', '未知文件名')
            file_size = image.get('file_size', '未知大小')
            mime_type = image.get('mime_type', '未知类型')
            image_time = image.get('time', '未知时间')

            if not url:
                logger.warning(f"Image URL missing for Execution ID {execution_id} at index {index}")
                continue

            # 添加图片和信息到 HTML
            html_content += f"""
            <div class="image-card">
                <div class="image-info">
                    <p><strong>图片 {index + 1} 信息:</strong></p>
                    <p>文件名: {file_name}</p>
                    <p>文件大小: {file_size} 字节</p>
                    <p>MIME 类型: {mime_type}</p>
                    <p>时间: {image_time}</p>
                    <p>URL: <a href="{url}" target="_blank">{url}</a></p>
                </div>
                <img src="{url}" alt="图片 {index + 1}">
            </div>
            """

        html_content += """
        </body>
        </html>
        """

        # 将 HTML 写入文件
        html_file = f"execution_{execution_id}_images.html"
        with open(html_file, "w", encoding="utf-8") as file:
            file.write(html_content)

        # 使用默认浏览器打开 HTML 文件
        html_path = os.path.abspath(html_file)
        webbrowser.open(f"file://{html_path}")
        logger.info(f"HTML file generated and opened in browser: {html_path}")

    def permutation_and_combination(self, event):
        """
        排列工具生成器
        """
        dlg = PermutationDialog(self, "排列生成工具")
        dlg.ShowModal()
        dlg.Destroy()

    def on_download(self, grid, event):
        """
        下载用例，包含图片数据
        :param grid:
        :param event:
        :return:
        """
        logger.info("----------------开始下载用例测试结果-----------------")
        # 获取Windows下载目录路径
        downloads_path = os.path.join(os.path.expanduser('~'), 'Downloads')
        # 设置文件名和路径
        now = datetime.now()
        formatted_now = now.strftime('%Y_%m_%d_%H_%M_%S')
        filename = f"TTS_result_{formatted_now}.xlsx"
        filepath = os.path.join(downloads_path, filename)

        # 创建一个Workbook和一个Worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # 设置列标题
        cols_title = ['测试结果', '测试耗时(S)', '用例标题', '前置条件',
                      '用例步骤', '预期结果', '开始时间', '完成时间', '评论', '失败次数', '阻塞次数', '图片数据']
        # 创建一个加粗的字体
        bold_font = Font(bold=True)
        ws.append(cols_title)
        for cell in ws["1:1"]:
            cell.font = bold_font

        # 填充数据
        for row in range(grid.GetNumberRows()):
            # 获取当前行的 ExecutionID
            execution_id = self.row_to_execution_id.get(row)
            image_data = ""

            # 如果 ExecutionID 存在，调用接口获取图片数据
            if execution_id:
                try:
                    response = http_manager.get_params(f'/get_images/{execution_id}')
                    images = response.get('images', [])
                    if images:
                        # 将图片的 URL 或 Base64 数据拼接为字符串
                        image_data = str(images)
                # 客观场景是不用上传图片的
                except BaseException as e:
                    logger.warning(f"下载用例图片时出现错误: {e}")

            # 获取当前行的所有单元格数据（排除第 0 列）
            row_data = [grid.GetCellValue(row, col) for col in range(1, grid.GetNumberCols() - 1)]
            # 将图片数据追加到行数据中
            row_data.append(image_data)
            # 写入 Excel 文件
            ws.append(row_data)

            # 根据测试结果设置单元格背景颜色
            if row_data[0] == 'Pass':
                for cell in ws[row + 2]:  # Excel行列都是从1开始计数，标题行占据第1行
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif row_data[0] == 'Fail' or row_data[0] == 'Block':
                for cell in ws[row + 2]:
                    cell.fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")

        # 保存文件
        try:
            wb.save(filepath)
            wx.MessageBox(f"文件已成功保存至 {filepath}", "保存成功", wx.OK | wx.ICON_INFORMATION)
        except IOError as e:
            wx.LogError(f"无法保存文件 '{filepath}'. 错误: {e}")

    def case_tree(self, sheet_id, model_id):
        """
        负责展示用例左侧节点
        :param sheet_id: 用例文件
        :return: 所有用例内容
        """
        logger.warning("开始展示用例树节点")
        self.tree.DeleteAllItems()  # 清空现有的树状结构
        self.testCases = http_manager.get_cases_by_sheet_id(sheet_id, model_id)  # 获取用例数据（列表形式）
        self.update_statistics()
        if self.tree.GetImageList() is None:
            self.tree.AssignImageList(self.image_list)

        # 添加根节点并设置图标
        root = self.tree.AddRoot(self.sheet_name, self.icon_indices['Root'])
        for case_data in self.testCases:
            # 解析用例数据
            case_id = case_data.get('CaseID')  # 用例ID
            case_status = case_data.get('TestResult')  # 用例状态
            case_title = case_data.get('CaseTitle')  # 用例标题

            # 根据状态添加相应的图标
            if case_status in self.icon_indices:
                case_node = self.tree.AppendItem(root, case_title)
                self.tree.SetItemImage(case_node, self.icon_indices[case_status])  # 设置节点图像
            else:
                case_node = self.tree.AppendItem(root, case_title)

            # 将CaseID存储在节点数据中
            self.tree.SetItemData(case_node, case_id)
        self.tree.Expand(root)

    def case_details(self, event):
        """
        负责展示用例详情
        :param event:
        :return:
        """
        logger.warning("开始展示用例详情")
        item = event.GetItem()
        parent = self.tree.GetItemParent(item)

        # 根节点不展示内容
        if self.tree.GetItemText(item) == self.sheet_name:
            self.content.SetValue("")
            self.CaseID = None
            return

        # 获取节点的CaseID
        self.CaseID = self.tree.GetItemData(item)
        logger.info(f"You selected the case with ID: {self.CaseID}")
        action_and_num = http_manager.get_params(f'/get_case_actions_and_num/{self.CaseID}').get('actions_and_num')
        self.actions_and_num.SetLabel(f"监控动作: {action_and_num}")

        # 在testCases中查找与CaseID匹配的用例
        case = next((c for c in self.testCases if c.get('CaseID') == self.CaseID), None)
        if case:
            self.content.Clear()
            self.log_content.Clear()
            self.content.SetValue(f"测试机型: {self.model_name}\n\n")
            self.content.AppendText(f"用例标题:\n{case.get('CaseTitle')}\n\n")
            self.content.AppendText(f"前置条件:\n{case.get('PreConditions')}\n\n")
            self.content.AppendText(f"操作步骤:\n{case.get('CaseSteps')}\n\n")
            self.content.SetInsertionPointEnd()  # 移动光标到末尾以便于添加蓝色文本
            self.content.SetDefaultStyle(wx.TextAttr(wx.BLUE))
            self.content.AppendText(f"预期结果:\n{case.get('ExpectedResult')}")
            # 再次将文本颜色设置回默认颜色，以防止后续文本也变为蓝色
            self.content.SetDefaultStyle(wx.TextAttr(wx.BLACK))
            wx.CallAfter(self.scroll_to_top)  # 调用滚动方法

    def scroll_to_top(self):
        """
        滚动到文本框的顶部
        """
        self.content.ScrollLines(-self.content.GetNumberOfLines())


class PermutationDialog(wx.Dialog):
    """
    排列工具
    """

    def __init__(self, parent, title):
        super().__init__(parent, title=title, size=(450, 300))

        panel = wx.Panel(self)
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        # 端口勾选框
        self.ports_checkboxes = []
        ports = ['C1', 'C2', 'C3', 'C4', 'A1', 'A2', 'A3', 'A4']
        ports_sizer = wx.BoxSizer(wx.HORIZONTAL)
        for port in ports:
            cb = wx.CheckBox(panel, label=port)
            self.ports_checkboxes.append(cb)
            ports_sizer.Add(cb, 1, wx.ALL, 5)

        main_sizer.Add(wx.StaticText(panel, label="请选择端口:"), 0, wx.ALL, 5)
        main_sizer.Add(ports_sizer, 0, wx.EXPAND | wx.ALL, 5)

        # 型号输入框
        main_sizer.Add(wx.StaticText(panel, label="请输入设备型号（每行一个）:"), 0, wx.ALL, 5)
        self.devices_textctrl = wx.TextCtrl(panel, style=wx.TE_MULTILINE, size=(300, 100))
        main_sizer.Add(self.devices_textctrl, 1, wx.EXPAND | wx.ALL, 5)

        # 生成按钮
        generate_button = wx.Button(panel, label="生成排列")
        generate_button.Bind(wx.EVT_BUTTON, self.on_generate)
        main_sizer.Add(generate_button, 0, wx.ALIGN_CENTER | wx.ALL, 10)

        panel.SetSizer(main_sizer)

    def on_generate(self, event):
        # 获取选中的端口
        selected_ports = [cb.GetLabel() for cb in self.ports_checkboxes if cb.IsChecked()]
        if not selected_ports:
            wx.MessageBox("请至少选择一个端口", "提示", wx.OK | wx.ICON_WARNING)
            return

        # 获取输入的设备型号
        devices = [device.strip() for device in self.devices_textctrl.GetValue().split('\n') if device.strip()]
        if len(devices) < len(selected_ports):
            wx.MessageBox("设备型号的数量不能小于选中的端口数量", "提示", wx.OK | wx.ICON_WARNING)
            return

        # 生成排列组合
        combinations = list(itertools.permutations(devices, len(selected_ports)))

        # 显示文件保存对话框，让用户选择保存路径和文件名
        with wx.FileDialog(self, "保存排列组合", wildcard="Excel files (*.xlsx)|*.xlsx",
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as file_dialog:

            if file_dialog.ShowModal() == wx.ID_CANCEL:
                return  # 用户取消操作

            # 获取用户选择的路径
            save_path = file_dialog.GetPath()
        # 创建 Excel 文件并保存
        try:
            self.save_to_excel(selected_ports, combinations, save_path)
            wx.MessageBox(f"排列组合已生成并保存为 {save_path}", "提示", wx.OK | wx.ICON_INFORMATION)
        except Exception as e:
            wx.MessageBox(f"保存文件时发生错误: {str(e)}", "错误", wx.OK | wx.ICON_ERROR)

    def save_to_excel(self, columns, data, file_path):
        # 创建 Excel 工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "排列组合"

        # 写入固定的用例模板标题
        ws.append(["测试机型", "用例标题", "前置条件", "用例步骤", "预期结果"])

        # 写入数据行
        for combination in data:
            ws.append(["", "[时间+1]" + ", ".join(columns), "", ", ".join(combination), "功能正常"])

        # 保存文件
        wb.save(file_path)
