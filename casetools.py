import pandas as pd
from openpyxl import load_workbook
from datetime import datetime


# 输入和输出文件的文件名

def case_tool(file_name):
    output_file_name = datetime.now().strftime("%Y%m%d") + '.xlsx'

    # 读取Excel文件中所有工作表的名称
    wb = load_workbook(file_name)
    sheets = wb.sheetnames
    wb.close()

    # 创建一个空的DataFrame来存储所有的数据
    all_data = pd.DataFrame()

    # 遍历每个sheet页，处理存在所需字段的数据
    for sheet in sheets:
        try:
            df = pd.read_excel(file_name, sheet_name=sheet)
            # 检查所需字段是否在DataFrame中
            required_columns_list = ['Section', 'Group', 'Case item', 'Expected result']
            if all(col in df.columns for col in required_columns_list):
                # 如果所需列存在，则选择并重命名这些列
                required_columns = df[required_columns_list]
                required_columns.columns = ['测试机型', '用例标题', '用例步骤', '预期结果']
                # 追加到all_data DataFrame中
                all_data = pd.concat([all_data, required_columns], ignore_index=True)
            else:
                print(f"工作表 {sheet} 中缺少所需字段，已被跳过。")
        except Exception as e:
            print(f"读取工作表 {sheet} 时出现异常：{e}")

    # 将所有数据写入新Excel文件的第一个工作表中
    all_data['执行结果'] = None
    all_data.to_excel(output_file_name, index=False)

    print(f"所有数据已经从 '{file_name}' 读取并合并到新的Excel文件 '{output_file_name}' 的第一个sheet中。")


if __name__ == '__main__':
    file_name = R'C:\Users\71958\Documents\键盘通用测试用例新模板20230920 中文版待修订 1.xlsx'

    case_tool(file_name)
