# -*- coding: utf-8 -*-
# 负责GUI界面展示以及交互逻辑
import os
import time

import wx
import wx.grid
import subprocess
import threading
import sqlite3
import openpyxl
from functools import partial

from openpyxl.styles import PatternFill

from common.rw_excel import MyExcel
from datetime import datetime
from common.logs import logger

class ResetButtonRenderer(wx.grid.GridCellRenderer):
    def __init__(self):
        wx.grid.GridCellRenderer.__init__(self)

    def Draw(self, grid, attr, dc, rect, row, col, isSelected):
        button_text = "重置"
        dc.SetBrush(wx.Brush("white"))
        dc.SetPen(wx.TRANSPARENT_PEN)
        dc.DrawRectangle(rect)
        tw, th = dc.GetTextExtent(button_text)
        # Centering text horizontally and vertically
        dc.DrawText(button_text, rect.x + rect.width // 2 - tw // 2, rect.y + rect.height // 2 - th // 2)

class TestCasesPanel(wx.Panel):
    def __init__(self, parent):
        super().__init__(parent)
        self.conn = sqlite3.connect('sqlite-tools/lenovoDb')  # 连接到数据库
        self.splitter = wx.SplitterWindow(self)

        self.tree = wx.TreeCtrl(self.splitter)
        self.content = wx.TextCtrl(self.splitter, style=wx.TE_MULTILINE | wx.TE_READONLY | wx.TE_RICH2)

        self.splitter.SplitVertically(self.tree, self.content)
        self.splitter.SetSashGravity(0.5)

        # 上传按钮
        upload_icon = wx.Bitmap("icon/上传文件夹.png")
        self.upload_button = wx.BitmapButton(self, bitmap=upload_icon)
        self.upload_button.SetToolTip(wx.ToolTip("上传文件"))
        self.upload_button.Bind(wx.EVT_BUTTON, self.up_file)

        # 检查按钮
        device_icon = wx.Bitmap("icon/设备管理.png")
        # self.check_button = wx.Button(self, label='检查')
        self.device_button = wx.BitmapButton(self, bitmap=device_icon)
        self.device_button.SetToolTip(wx.ToolTip("打开设备管理器"))
        self.device_button.Bind(wx.EVT_BUTTON, self.open_device_manager)

        # 配置按钮
        config_icon = wx.Bitmap("icon/configure.png")
        self.config_button = wx.BitmapButton(self, bitmap=config_icon)
        self.config_button.SetToolTip(wx.ToolTip("生成配置文件"))
        self.config_button.Bind(wx.EVT_BUTTON, self.write_config)

        # 查看用例状态按钮
        status_icon = wx.Bitmap("icon/查看状态1.png")
        self.status_button = wx.BitmapButton(self, bitmap=status_icon)
        self.status_button.SetToolTip(wx.ToolTip("查看用例状态"))
        self.status_button.Bind(wx.EVT_BUTTON, self.check_status)

        # 附件按钮
        annex_icon = wx.Bitmap("icon/附件.png")
        self.annex_button = wx.BitmapButton(self, bitmap=annex_icon)
        self.annex_button.SetToolTip(wx.ToolTip("上传附件"))
        self.annex_button.Bind(wx.EVT_BUTTON, self.select_annex)

        # 图片
        lenovo_bitmap = wx.Bitmap('icon/3332.png', wx.BITMAP_TYPE_ANY)
        imageCtrl = wx.StaticBitmap(self, bitmap=lenovo_bitmap)

        # 下拉框 监控动作、测试次数
        self.monitor_actions = ['S3', 'S4', 'S5', 'Restart']
        self.tests_num = ['1', '3', '5', '10', '20', '50', '100']
        self.actions_box = wx.ComboBox(self, choices=self.monitor_actions)
        self.tests_box = wx.ComboBox(self, choices=self.tests_num)

        # start、pass、fail、block 按钮
        self.start_button = wx.Button(self, label='start')
        self.start_button.Bind(wx.EVT_BUTTON, self.start_test)

        # 创建布局，将按钮放置在左上角
        buttonSizer = wx.BoxSizer(wx.HORIZONTAL)
        mainSizer = wx.BoxSizer(wx.VERTICAL)
        imgSizer = wx.BoxSizer(wx.HORIZONTAL)

        # 将按钮添加到布局中
        buttonSizer.Add(self.upload_button, 0, wx.ALL, 5)
        buttonSizer.Add(self.device_button, 0, wx.ALL, 5)
        buttonSizer.Add(self.config_button, 0, wx.ALL, 5)
        buttonSizer.Add(self.status_button, 0, wx.ALL, 5)
        buttonSizer.Add(self.annex_button, 0, wx.ALL, 5)

        # 创建新的布局，将下拉框和按钮放在右下脚
        actionSizer = wx.BoxSizer(wx.HORIZONTAL)
        buttonSizer2 = wx.BoxSizer(wx.HORIZONTAL)
        # 创建按钮并添加到布局
        self.result_buttons = {}
        for button in ['Pass', 'Fail', 'Block']:
            self.result_buttons[button] = wx.Button(self, label=button)
            self.result_buttons[button].Bind(wx.EVT_BUTTON, self.test_result)
            buttonSizer2.Add(self.result_buttons[button], 0, wx.ALL, 5)
            self.result_buttons[button].Hide()

        # 添加下拉框到新的布局
        actionSizer.Add(wx.StaticText(self, label="监控动作:"), 0, wx.ALL, 5)
        actionSizer.Add(self.actions_box, 0, wx.ALL, 5)
        actionSizer.Add(wx.StaticText(self, label="测试次数:"), 0, wx.ALL, 5)
        actionSizer.Add(self.tests_box, 0, wx.ALL, 5)

        # 添加按钮到新的布局
        buttonSizer2.Add(self.start_button, 0, wx.ALL, 5)
        # buttonSizer2.Add(self.pass_button, 0, wx.ALL, 5)
        # buttonSizer2.Add(self.fail_button, 0, wx.ALL, 5)
        # buttonSizer2.Add(self.block_button, 0, wx.ALL, 5)



        # 在主布局中添加按钮布局和新的布局
        mainSizer.Add(buttonSizer, 0, wx.ALIGN_LEFT)
        mainSizer.Add(self.splitter, 1, wx.EXPAND)


        # 创建底部布局，将下拉框和按钮放在右侧
        bottomSizer = wx.BoxSizer(wx.HORIZONTAL)
        # 添加图片到布局（最左边）
        bottomSizer.Add(imageCtrl, 0, wx.ALL | wx.ALIGN_LEFT, 5)
        bottomSizer.Add(actionSizer, 0, wx.ALL, 5)
        bottomSizer.Add(buttonSizer2, 0, wx.ALL, 5)

        mainSizer.Add(bottomSizer, 0, wx.ALL, 5)

        # 设置主布局
        self.SetSizer(mainSizer)
        self.testCases = self.PopulateTree()
        # 用户点击标题赋值
        self.CaseID = None
        self.tree.Bind(wx.EVT_TREE_SEL_CHANGED, self.OnSelChanged)



    def test_result(self, event):
        clicked_button = event.GetEventObject()
        if clicked_button is self.result_buttons['Pass']:
            # 处理 pass
            logger.info(f"Pass Button clicked {self.CaseID}")
            wx.CallAfter(self.case_enable)
        elif clicked_button in [self.result_buttons['Fail'], self.result_buttons['Block']]:
            # 处理 fail or block
            # 弹出文本输入对话框
            dlg = wx.TextEntryDialog(self, '请输入Fail/Block原因 :', 'Comment')
            if dlg.ShowModal() == wx.ID_OK:
                input_content = dlg.GetValue().strip()  # 获取输入的内容
                if input_content:
                    logger.info(f"{clicked_button.GetLabel()} Button clicked, Content: {input_content}")
                    wx.CallAfter(self.case_enable)
                else:
                    wx.MessageDialog(self, '内容不能为空，请重新输入!', '错误', style=wx.OK | wx.ICON_ERROR).ShowModal()
            dlg.Destroy()


    def start_test(self, event):
        action = self.actions_box.GetValue()
        num_test = self.tests_box.GetValue()
        if not action or not num_test:
            wx.MessageBox('监控动作/测试次数不能为空', 'Warning')
            return
        wx.CallAfter(self.case_disable)
        # 使用多线程异步运行，防止GUI界面卡死
        if action == 'S3':
            thread = threading.Thread(target=self.test_S3, args=(int(num_test),))
            thread.start()
        elif action == 'S4':
            pass  # Handle other actions

    def test_S3(self, num_test):
        count = 0
        while count < num_test:
            count += 1
            time.sleep(1)
            print(f"Running S3 test {count} of {num_test}")
        wx.CallAfter(self.after_test)

    def after_test(self):
        for button in self.result_buttons:
            self.result_buttons[button].Enable()
        self.Layout()

    def case_disable(self):
        """
        case 执行状态按钮显示
        :return:
        """
        self.tests_box.Disable()
        self.actions_box.Disable()
        self.tree.Disable()
        self.start_button.Hide()
        for button in self.result_buttons:
            self.result_buttons[button].Show()
            self.result_buttons[button].Disable()
        self.Layout()

    def case_enable(self):
        """
        case 未执行状态按钮显示
        :return:
        """
        self.tree.Enable()
        self.actions_box.Enable()
        self.tests_box.Enable()
        self.start_button.Show()
        for button in self.result_buttons:
            self.result_buttons[button].Hide()
        self.Layout()

    def test_S4(self):
        # Implement the steps for the test S4 function here
        pass

    def up_file(self, event):
        # 打开文件对话框
        with wx.FileDialog(self, "选择测试用例文件", wildcard="文本文件 (*.xlsx)|*.xlsx",
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return  # 用户取消了操作

            # 获取选择的文件路径/读取用例
            pathname = fileDialog.GetPath()
            file_name = os.path.basename(pathname)
            data = MyExcel(pathname)
            data.active_sheet('Sheet1')
            col_data = data.getRowValues(1)
            case_data = data.get_appoint_row_values(2)
            logger.info(f'excel case data is {case_data}')
            cur = self.conn.cursor()
            try:
                # 校验格式
                data.validate_case_data(col_data)
                # 检查 file_name 是否已经存在
                cur.execute(f"SELECT COUNT(*) FROM TestFile WHERE FileName = '{file_name}'")
                if cur.fetchone()[0] == 0:  # 统计结果为0插入到 TestFile 表
                    logger.info(f"start inserting {file_name} into DB ")
                    cur.execute(f"INSERT INTO TestFile (FileName) VALUES ('{file_name}')")
                    # 获取刚刚插入的 TestFile 记录的 FileID
                    file_id = cur.lastrowid
                    now = datetime.now()
                    formatted_now = now.strftime('%Y-%m-%d %H:%M:%S')
                    for case in case_data:
                        cur.execute(f"INSERT INTO TestCase (ModelName, CaseTitle, CaseSteps, ExpectedResult, TestResult, ExecutionTime, FileID) \
                                        VALUES ('{case[0]}', '{case[1]}', '{case[2]}', '{case[3]}', '{case[4]}', '{formatted_now}', {file_id})")
                else:
                    wx.MessageBox(f"文件 '{file_name}' 已经存在于数据库中。", "提示", wx.OK | wx.ICON_INFORMATION)
            except ValueError as e:
                # 格式校验出错
                wx.MessageBox(f"上传的用例格式不符合规则，请按照模板导入用例。错误详情: {e}", "提示",
                              wx.OK | wx.ICON_WARNING)
            except Exception as e:
                # 出现异常回滚
                self.conn.rollback()
                logger.info(f"An error occurred: {e}")
            else:
                self.conn.commit()
                self.testCases = self.PopulateTree(file_name)
                # 使用 partial可以提前填充一个参数，得到一个只需要一个参数的新函数
                self.tree.Bind(wx.EVT_TREE_SEL_CHANGED, partial(self.OnSelChanged, filename=file_name))
            finally:
                cur.close()


    def open_device_manager(self, event):
        """
        打开Windows设备管理器
        :param event:
        :return:
        """

        def run_command():
            subprocess.run('mmc devmgmt.msc', shell=True)

        threading.Thread(target=run_command).start()

    def write_config(self, event):
        """
        写入配置信息/暂时不知道具体用途
        :param event:
        :return:
        """
        # 创建一个对话框用于输入配置信息
        dialog = wx.Dialog(None, title="配置信息", size=(350, 350))
        panel = wx.Panel(dialog)
        vbox = wx.BoxSizer(wx.VERTICAL)

        # 创建 BIOS,ECFW,MEFW, System SN,Retimer 的输入项
        labels = ["BIOS", "ECFW", "MEFW", "System SN", "Retimer"]
        self.inputs = {}
        for label in labels:
            hbox = wx.BoxSizer(wx.HORIZONTAL)
            st = wx.StaticText(panel, label=label)
            tc = wx.TextCtrl(panel)
            hbox.Add(st)
            hbox.Add(tc, flag=wx.LEFT, border=5)
            vbox.Add(hbox, flag=wx.EXPAND | wx.ALL, border=10)
            self.inputs[label] = tc

        # 为确认和取消按钮创建一行
        hbox = wx.BoxSizer(wx.HORIZONTAL)
        okButton = wx.Button(panel, label='保存')
        closeButton = wx.Button(panel, label='取消')

        # 绑定按钮事件
        okButton.Bind(wx.EVT_BUTTON, partial(self.save_config, dialog=dialog))
        closeButton.Bind(wx.EVT_BUTTON, partial(self.close_dialog, dialog=dialog))

        hbox.Add(okButton)
        hbox.Add(closeButton, flag=wx.LEFT, border=5)
        vbox.Add(hbox, flag=wx.ALIGN_CENTER | wx.TOP | wx.BOTTOM, border=10)

        panel.SetSizer(vbox)
        dialog.ShowModal()
        dialog.Destroy()

    def save_config(self, event, dialog):
        with open('config.txt', 'w') as f:
            for label, tc in self.inputs.items():
                f.write(f"{label}: {tc.GetValue()}\n")
        dialog.Close()
        wx.MessageBox("配置文件保存成功！", "信息", wx.OK | wx.ICON_INFORMATION)

    def close_dialog(self, event, dialog):
        dialog.Close()

    def check_status(self, event):
        # 创建一个新的对话框，并且允许对话框最大化
        dialog = wx.Dialog(self, title="查看用例状态", style=wx.MAXIMIZE_BOX | wx.DEFAULT_DIALOG_STYLE)

        # 创建grid并设置行和列
        grid = wx.grid.Grid(dialog)
        grid.CreateGrid(numRows=len(self.testCases), numCols=len(self.testCases[0])-2)

        # 设置列标题
        cols_title = ['测试结果', '测试机型', '用例标题', '前置条件', '用例步骤', '预期结果', '执行时间', '完成时间']
        for i, title in enumerate(cols_title):
            grid.SetColLabelValue(i, title)

        # 填充数据
        for i, case in enumerate(self.testCases):
            for j, item in enumerate(case[:-2]):  # 排除ID等敏感数据
                grid.SetCellValue(i, j, str(item))  # 第i行，第j列，数据
                # 检查测试结果列，设置背景颜色
                if cols_title[j] == '测试结果':
                    if item == 'pass':
                        grid.SetCellBackgroundColour(i, j, wx.Colour(144, 238, 144))  # 浅绿色
                    elif item == 'fail':
                        grid.SetCellBackgroundColour(i, j, wx.Colour(255, 99, 71))  # 浅红色

        # 自动调整每一列和每一行的大小以适应内容
        grid.AutoSizeColumns()
        grid.AutoSizeRows()
        # 添加重置按钮
        grid.InsertCols(0)
        cols_title.insert(0, "重置按钮")
        for i in range(len(self.testCases)):
            grid.SetCellRenderer(i, 0, ResetButtonRenderer())

        grid.Bind(wx.grid.EVT_GRID_CELL_LEFT_CLICK, self.on_reset_click)

        # 创建下载按钮
        download_button = wx.Button(dialog, label="下载")
        download_button.Bind(wx.EVT_BUTTON, lambda evt: self.on_download(grid, evt))

        # 创建对话框的布局管理器，并将grid添加到其中
        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(grid, 1, wx.EXPAND | wx.ALL)
        sizer.Add(download_button, 0, wx.ALIGN_CENTER | wx.ALL, 10)
        dialog.SetSizer(sizer)

        # 显示对话框
        dialog.ShowModal()
        dialog.Destroy()

    def on_reset_click(self, evt):
        row, col = evt.GetRow(), evt.GetCol()
        # Check if "Reset" cell has been clicked
        if col == 0:
            reset_msg = wx.MessageDialog(self, '您确定要重置这条用例测试结果吗?', '确认', wx.YES_NO | wx.ICON_QUESTION)
            reset_response = reset_msg.ShowModal()
            if reset_response == wx.ID_YES:
                # TODO: Reset the appropriate test case result
                # For example: self.testCases[row][0] = 'pending'
                logger.info('222222222222222222222222222222222222222')
                evt.Skip(False)
            reset_msg.Destroy()
            return
        evt.Skip(True)
    def select_annex(self, event):
        pass

    def on_download(self, grid, event):
        """
        下载用例
        :param grid:
        :param event:
        :return:
        """
        logger.info("----------------执行下载测试结果-----------------")
        # 获取Windows下载目录路径
        downloads_path = os.path.join(os.path.expanduser('~'), 'Downloads')
        # 设置文件名和路径
        now = datetime.now()
        formatted_now = now.strftime('%Y_%m_%d_%H_%M_%S')
        filename = f"PATVS_result_{formatted_now}.xlsx"
        filepath = os.path.join(downloads_path, filename)

        # 创建一个Workbook和一个Worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # 设置列标题
        cols_title = ['测试结果', '测试机型', '用例标题', '前置条件', '用例步骤', '预期结果', '执行时间', '完成时间']
        ws.append(cols_title)
        # 填充数据
        for row in range(grid.GetNumberRows()):
            row_data = [grid.GetCellValue(row, col) for col in range(1, grid.GetNumberCols())]
            ws.append(row_data)

            # 根据测试结果设置单元格背景颜色
            if row_data[0] == 'pass':
                for cell in ws[row + 2]:  # Excel行列都是从1开始计数，标题行占据第1行
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif row_data[0] == 'fail':
                for cell in ws[row + 2]:
                    cell.fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")

        # 保存文件
        try:
            wb.save(filepath)
            wx.MessageBox(f"文件已成功保存至 {filepath}", "保存成功", wx.OK | wx.ICON_INFORMATION)
        except IOError as e:
            wx.LogError(f"无法保存文件 '{filepath}'. 错误: {e}")

    def PopulateTree(self, filename='Demo测试用例'):
        """
        负责展示用例左侧节点
        :param filename: 用例文件
        :return: 所有用例内容
        """
        self.tree.DeleteAllItems()  # 清空现有的树状结构
        root = self.tree.AddRoot(filename)
        cur = self.conn.cursor()
        cur.execute(f"SELECT FileID FROM TestFile where FileName = '{filename}'")
        result = cur.fetchone()
        logger.info(f'filename is {result}')
        if result is not None:
            file_id = result[0]
        else:
            # 为空展示 demo 文件
            file_id = 1
        # cur.execute(f'SELECT ModelName, CaseTitle, CaseSteps, ExpectedResult FROM TestCase where FileID = {file_id}')
        cur.execute(f'SELECT * FROM TestCase where FileID = {file_id}')
        all_case = cur.fetchall()
        logger.info(f"{filename} all case is {all_case}")
        for row in all_case:
            caseID = row[8]
            caseTitle = row[2]
            caseSteps = row[4]
            expectedResult = row[5]
            caseNode = self.tree.AppendItem(root, caseTitle)
            self.tree.SetItemData(caseNode, caseID)
            self.tree.AppendItem(caseNode, f"操作步骤: {caseSteps}")
            self.tree.SetItemData(caseNode, caseID)
            self.tree.AppendItem(caseNode, f"预期结果: {expectedResult}")
            self.tree.SetItemData(caseNode, caseID)
        self.tree.Expand(root)
        cur.close()
        return all_case

    def on_tree_selection_changed(self, event):
        item = event.GetItem()
        item_data = self.tree.GetItemData(item)
        if item_data is not None:
            logger.info(f"You selected the case with ID: {item_data}")

    def OnSelChanged(self, event, filename='Demo测试用例'):
        """
        负责展示用例详情
        :param event:
        :return:
        """
        item = event.GetItem()
        parent = self.tree.GetItemParent(item)

        # 根节点不展示内容
        if self.tree.GetItemText(item) == filename:
            self.content.SetValue("")
            self.CaseID = None
            return
        # 获取用例的根节点，以便于处理多层子节点
        while self.tree.GetItemParent(parent).IsOk():
            item = parent
            parent = self.tree.GetItemParent(item)

        caseTitle = self.tree.GetItemText(item)
        self.CaseID = self.tree.GetItemData(item)
        logger.info(f"You selected the case with ID: {self.CaseID}")
        for case in self.testCases:
            if case[2] == caseTitle:
                self.content.Clear()
                self.content.SetValue(f"用例标题:\n{case[2]}\n\n")
                self.content.AppendText(f"前置条件:\n{case[3]}\n\n")
                self.content.AppendText(f"操作步骤:\n{case[4]}\n\n")
                self.content.SetInsertionPointEnd()  # 移动光标到末尾以便于添加蓝色文本
                self.content.SetDefaultStyle(wx.TextAttr(wx.BLUE))
                self.content.AppendText(f"预期结果:\n{case[5]}")
                # 再次将文本颜色设置回默认颜色，以防止后续文本也变为蓝色
                self.content.SetDefaultStyle(wx.TextAttr(wx.BLACK))
                break


app = wx.App(False)
frame = wx.Frame(None, title="PATVS-v1.0.0", size=(800, 600))
panel = TestCasesPanel(frame)
frame.Show()
app.MainLoop()
